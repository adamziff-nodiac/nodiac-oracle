# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl", "requests", "shapely", "gridstatus", "pandas"]
# ///
"""
Real county scores data pipeline.

Downloads and processes public datasets to compute per-county scores
for the Regional Hub Strategy heat map.

Usage:
    uv run scripts/build-real-county-scores.py

Data Sources:
    - Co-op Density:     EIA Form 861 (2024) — Service Territory + Frame
    - Grid Reliability:  EIA Form 861 (2024) — Reliability metrics
    - Curtailment Proxy: EIA Form 860 (2024) — Variable renewable MW per county
    - Labor:             Census County Business Patterns (2023) — NAICS 5182/5415/517
    - Fiber:             FCC BDC Dec 2024 (primary) — Fiber availability at BSLs via ArcGIS
                         Census ACS 5-Year 2023 (fallback) — Broadband subscriptions proxy
                         NTIA Middle Mile grants (CFDA 11.033) — Middle-mile fiber infrastructure
    - Permitting:        State DC incentive data (NCSL, SDI Alliance, H5, NAIOP, Data Center Watch)

Output:
    - public/data/county-scores.json (static fallback for frontend)
    - Optional: Supabase upsert (if SUPABASE_URL and SUPABASE_SECRET_KEY set)
"""

import csv
import io
import json
import math
import os
import re
import sys
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl
import requests
from shapely.geometry import shape, mapping
from shapely.ops import unary_union
from shapely.prepared import prep
from shapely import STRtree

# ============================================================
# Configuration
# ============================================================

# County FIPS reference from Census
FIPS_URL = "https://www2.census.gov/geo/docs/reference/codes2020/national_county2020.txt"

# EIA Form 861 (co-op density + grid reliability)
EIA_861_URL = "https://www.eia.gov/electricity/data/eia861/zip/f8612024.zip"

# EIA Form 861 multi-year reliability data (2013-2024, reliability reporting started 2013)
EIA_861_RELIABILITY_YEARS = list(range(2013, 2025))  # 2013 through 2024
EIA_861_ARCHIVE_URL_TEMPLATE = "https://www.eia.gov/electricity/data/eia861/archive/zip/f861{year}.zip"
EIA_861_CURRENT_URL = "https://www.eia.gov/electricity/data/eia861/zip/f8612024.zip"

# EIA Form 860 (renewable capacity / curtailment proxy)
EIA_860_URL = "https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip"

# EIA Form 923 (actual generation data for capacity factor gap analysis)
EIA_923_URL = "https://www.eia.gov/electricity/data/eia923/xls/f923_2024.zip"

# Census CBP (IT labor)
CBP_API_BASE = "https://api.census.gov/data/2023/cbp"
CBP_NAICS_CODES = ["5182", "5415", "517"]

# Census population estimates
POP_URL = "https://www2.census.gov/programs-surveys/popest/datasets/2020-2024/counties/totals/co-est2024-alldata.csv"

# Census ACS 5-Year broadband (fiber fallback)
ACS_BROADBAND_URL = "https://api.census.gov/data/2023/acs/acs5"

# FCC BDC (primary fiber source) — via ArcGIS Living Atlas county summaries
# Data: FCC Broadband Data Collection, Dec 2024 vintage
# Fields: TotalBSLs, ServedBSLsFiber, UnderservedBSLsFiber, UniqueProvidersFiber
FCC_BDC_ARCGIS_URL = (
    "https://services8.arcgis.com/peDZJliSvYims39Q/arcgis/rest/services/"
    "FCC_Broadband_Data_Collection_December_2024_View/FeatureServer/1/query"
)

# ArcGIS co-op + public power service territory polygons (Oak Ridge / LANL / INL)
COOP_TERRITORY_ARCGIS_URL = (
    "https://services5.arcgis.com/ARxOqVFcodl7rmzw/arcgis/rest/services/"
    "America_Electrical_Coop_Service_Territories/FeatureServer/10/query"
)

# LBNL "Queued Up" interconnection queue data (2025 Edition, through 2024)
LBNL_QUEUE_URL = (
    "https://eta-publications.lbl.gov/sites/default/files/2025-08/"
    "lbnl_ix_queue_data_file_thru2024_v2.xlsx"
)

# FIPS crosswalk for EIA data
FIPS_CROSSWALK_URL = "https://raw.githubusercontent.com/kjhealy/fips-codes/master/state_and_county_fips_master.csv"

# Skip US territories
SKIP_STATES = {"AS", "GU", "MP", "PR", "VI"}

# Variable renewable technologies (for curtailment score)
VARIABLE_RENEWABLES = {
    "Solar Photovoltaic",
    "Solar Thermal with Energy Storage",
    "Solar Thermal without Energy Storage",
    "Onshore Wind Turbine",
    "Offshore Wind Turbine",
}

# ISO/RTO curtailment intensity scores (0-1 scale)
# Based on 2023-2024 published market reports and curtailment data
# Higher = more curtailment observed in this balancing authority
ISO_CURTAILMENT_INTENSITY = {
    # ERCOT: 8+ TWh curtailed in 2024, worst in US
    # West Zone + Panhandle account for majority
    # Sources: Modo Energy (2024), World Climate Service (Oct 2024)
    "ERCO": 0.95,

    # CAISO: 3.4M MWh curtailed in 2024, up 29% from 2023
    # 93% solar, concentrated in SP15 during midday shoulder seasons
    # Sources: EIA Today in Energy (2024), CAISO Production Data, Utility Dive
    "CISO": 0.85,

    # SPP: Average hourly wind curtailment 1,097 MW (2023), 6x increase since 2020
    # West and Panhandle zones generate more than transmission can export
    # Source: SPP 2024 Annual State of the Market Report
    "SWPP": 0.80,

    # MISO: Average hourly wind curtailment 508 MW (2023), up from 242 MW (2019)
    # West Region most congested
    # Source: MISO/Potomac Economics 2024 State of Market Report
    "MISO": 0.60,

    # BPA: Hydro/wind interaction causes curtailment during high-water seasons
    # Pacific NW wind curtailment during spring runoff
    # Source: BPA oversupply management reports
    "BPAT": 0.55,

    # Imperial Irrigation District: Moderate solar curtailment
    # Small but congested territory in Southern California desert
    "IID": 0.50,

    # PJM: Curtailments jumped 6x in 2024 vs 2023
    # Driven by Northern Virginia data center congestion
    # Source: PJM 2025 Renewable Dispatch Data Request Results
    "PJM": 0.40,

    # Nevada Power: Moderate solar curtailment
    # Growing with Nevada solar buildout
    "NEVP": 0.35,

    # NYISO: Moderate, growing with offshore wind pipeline
    "NYIS": 0.25,

    # ISO-NE: Lower curtailment, but growing
    "ISNE": 0.20,
}
# Default for balancing authorities not listed above
ISO_CURTAILMENT_DEFAULT = 0.10

# State-level expected capacity factors for solar PV and wind (2022-2024 averages)
# Sources: EIA Electric Power Monthly Table 6.07.B, NREL ATB 2024
# Used as benchmarks to detect curtailment via CF gap analysis
EXPECTED_CF_SOLAR = {
    "AZ": 0.27, "CA": 0.27, "NV": 0.26, "NM": 0.26, "UT": 0.25,
    "TX": 0.24, "CO": 0.24, "FL": 0.23, "NC": 0.22, "GA": 0.22,
    "SC": 0.22, "AL": 0.21, "MS": 0.21, "LA": 0.21, "TN": 0.21,
    "OK": 0.21, "KS": 0.21, "AR": 0.20, "MO": 0.20, "VA": 0.20,
    "KY": 0.19, "IN": 0.19, "IL": 0.19, "OH": 0.18, "PA": 0.18,
    "NJ": 0.18, "MD": 0.18, "DE": 0.18, "IA": 0.18, "NE": 0.19,
    "SD": 0.19, "ND": 0.18, "MN": 0.18, "WI": 0.17, "MI": 0.17,
    "NY": 0.17, "CT": 0.17, "MA": 0.17, "RI": 0.17, "VT": 0.16,
    "NH": 0.16, "ME": 0.16, "WV": 0.18, "ID": 0.22, "MT": 0.20,
    "WY": 0.22, "OR": 0.20, "WA": 0.18, "HI": 0.22, "AK": 0.12,
}
EXPECTED_CF_WIND = {
    "TX": 0.34, "OK": 0.38, "KS": 0.40, "NE": 0.38, "SD": 0.42,
    "ND": 0.40, "MN": 0.35, "IA": 0.36, "MO": 0.30, "IL": 0.30,
    "IN": 0.28, "OH": 0.26, "MI": 0.28, "WI": 0.28, "CO": 0.33,
    "WY": 0.38, "MT": 0.35, "NM": 0.32, "AZ": 0.25, "CA": 0.28,
    "OR": 0.28, "WA": 0.28, "ID": 0.28, "NV": 0.22, "UT": 0.25,
    "NY": 0.25, "PA": 0.24, "ME": 0.30, "VT": 0.26, "NH": 0.25,
    "MA": 0.28, "RI": 0.28, "CT": 0.25, "NJ": 0.25, "MD": 0.25,
    "VA": 0.26, "NC": 0.26, "WV": 0.26, "AR": 0.28, "LA": 0.25,
    "AL": 0.25, "GA": 0.25, "SC": 0.25, "FL": 0.22, "MS": 0.25,
    "TN": 0.25, "KY": 0.25, "DE": 0.25, "HI": 0.30, "AK": 0.28,
}

# NTIA Enabling Middle Mile Broadband Infrastructure Program (CFDA 11.033)
# USAspending.gov API for federal grant data
USASPENDING_SEARCH_URL = "https://api.usaspending.gov/api/v2/search/spending_by_award/"
USASPENDING_AWARD_URL = "https://api.usaspending.gov/api/v2/awards/"
NTIA_MIDDLE_MILE_CFDA = "11.033"

# State name -> abbreviation mapping (for parsing grant descriptions)
STATE_NAME_TO_ABBR: dict[str, str] = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
    "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID",
    "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS",
    "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME", "MARYLAND": "MD",
    "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE",
    "NEVADA": "NV", "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ",
    "NEW MEXICO": "NM", "NEW YORK": "NY", "NORTH CAROLINA": "NC",
    "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC",
    "SOUTH DAKOTA": "SD", "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT",
    "VERMONT": "VT", "VIRGINIA": "VA", "WASHINGTON": "WA",
    "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

# Output path
OUTPUT_PATH = Path("public/data/county-scores.json")


# ============================================================
# Utility functions
# ============================================================

def log(msg: str):
    print(f"  {msg}", flush=True)


def download(url: str, desc: str) -> bytes:
    """Download a URL with progress indication."""
    print(f"\n>>> Downloading {desc}...", flush=True)
    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    size_mb = len(resp.content) / 1024 / 1024
    log(f"Downloaded {size_mb:.1f} MB")
    return resp.content


def percentile_rank(values: list[float]) -> list[float]:
    """Convert values to percentile ranks (0-1)."""
    n = len(values)
    if n == 0:
        return []
    indexed = sorted(enumerate(values), key=lambda x: x[1])
    ranks = [0.0] * n
    for rank, (orig_idx, _) in enumerate(indexed):
        ranks[orig_idx] = rank / max(n - 1, 1)
    return ranks


def normalize_county_name(name: str) -> str:
    """Normalize county name for FIPS matching."""
    name = name.strip()
    # Fix double spaces
    while "  " in name:
        name = name.replace("  ", " ")
    # St -> St.
    if name.startswith("St "):
        name = "St. " + name[3:]
    # Ste -> Ste.
    if name.startswith("Ste "):
        name = "Ste. " + name[4:]
    return name


# Manual FIPS overrides for known mismatches
MANUAL_FIPS: dict[tuple[str, str], str] = {
    ("FL", "Miami Dade"): "12086",
    ("FL", "miami dade"): "12086",
    ("MD", "Prince Georges"): "24033",
    ("MD", "prince georges"): "24033",
    ("MD", "Queen Annes"): "24035",
    ("MD", "queen annes"): "24035",
    ("LA", "DeSoto"): "22031",
    ("LA", "desoto"): "22031",
    ("IL", "DeWitt"): "17039",
    ("IL", "dewitt"): "17039",
    ("NM", "Dona Ana"): "35013",
    ("NM", "dona ana"): "35013",
    ("SD", "Oglala Lakota"): "46102",
    ("SD", "oglala lakota"): "46102",
    # Alaska census areas (old names -> current FIPS)
    ("AK", "Juneau"): "02110",
    ("AK", "juneau"): "02110",
    ("AK", "Sitka"): "02220",
    ("AK", "sitka"): "02220",
    ("AK", "Yakutat"): "02282",
    ("AK", "yakutat"): "02282",
    ("AK", "Kusilvak"): "02158",
    ("AK", "kusilvak"): "02158",
    ("AK", "Matanuska Susitna"): "02170",
    ("AK", "matanuska susitna"): "02170",
    ("AK", "Prince of Wales Ketchikan"): "02198",
    ("AK", "prince of wales ketchikan"): "02198",
    ("AK", "Skagway Hoonah Angoon"): "02105",
    ("AK", "skagway hoonah angoon"): "02105",
    ("AK", "Valdez Cordova"): "02261",
    ("AK", "valdez cordova"): "02261",
    ("AK", "Wrangell Petersburg"): "02275",
    ("AK", "wrangell petersburg"): "02275",
    ("AK", "Yukon Koyukuk"): "02290",
    ("AK", "yukon koyukuk"): "02290",
    # St. Mary's MD special case
    ("MD", "St Marys"): "24037",
    ("MD", "st marys"): "24037",
    # Louisiana double-space fix
    ("LA", "St  Helena"): "22091",
    ("LA", "st  helena"): "22091",
}


# ============================================================
# Step 1: Build FIPS lookup
# ============================================================

def build_fips_lookup() -> dict[tuple[str, str], str]:
    """Build (state_abbr, county_name_lower) -> 5-digit FIPS lookup."""
    print("\n=== Building FIPS lookup ===", flush=True)

    data = download(FIPS_CROSSWALK_URL, "FIPS crosswalk")
    reader = csv.DictReader(io.StringIO(data.decode("utf-8")))

    lookup: dict[tuple[str, str], str] = {}
    for row in reader:
        state = row["state"]
        name = row["name"]
        fips = row["fips"].zfill(5)
        if state == "NA":
            continue

        # Strip common suffixes
        for suffix in [" County", " Parish", " Borough", " Census Area",
                       " Municipality", " city"]:
            name = name.replace(suffix, "")
        name = name.strip()

        lookup[(state, name.lower())] = fips

    log(f"FIPS lookup: {len(lookup)} entries")
    return lookup


def resolve_fips(state: str, county: str, fips_lookup: dict) -> str | None:
    """Resolve (state, county_name) to 5-digit FIPS code."""
    # Try manual override first
    manual = MANUAL_FIPS.get((state, county))
    if manual:
        return manual
    manual = MANUAL_FIPS.get((state, county.lower()))
    if manual:
        return manual

    # Normalize and try
    normalized = normalize_county_name(county)
    fips = fips_lookup.get((state, normalized.lower()))
    if fips:
        return fips

    # Try original (no normalization)
    fips = fips_lookup.get((state, county.lower().strip()))
    if fips:
        return fips

    return None


# ============================================================
# Step 2: Co-op Density from EIA-861
# ============================================================

def compute_coop_density_area(fips_lookup: dict) -> dict[str, float]:
    """Compute co-op density as % of county area covered by co-op/public power territories.

    Uses ArcGIS "America Electrical Coop Service Territories" layer (833 polygons,
    sourced from Oak Ridge / LANL / INL) intersected with Census county boundaries.

    Returns {fips_code: float(0-1)} where 1.0 = entire county covered.
    """
    print("\n=== Computing Co-op Density (Area-Based, ArcGIS) ===", flush=True)

    # Step 1: Download co-op territory polygons from ArcGIS
    log("Downloading co-op/public power territory polygons from ArcGIS...")
    all_features = []
    offset = 0
    batch_size = 1000  # 833 total, so one batch should suffice

    while True:
        params = {
            "where": "1=1",
            "outFields": "NAME,STATE",
            "outSR": "4326",
            "returnGeometry": "true",
            "resultRecordCount": str(batch_size),
            "resultOffset": str(offset),
            "f": "json",
        }
        resp = requests.get(COOP_TERRITORY_ARCGIS_URL, params=params, timeout=120)
        resp.raise_for_status()
        data = resp.json()
        features = data.get("features", [])
        if not features:
            break
        all_features.extend(features)
        log(f"  Downloaded {len(all_features)} features so far...")
        if len(features) < batch_size:
            break
        offset += batch_size

    log(f"Total co-op/public power territories: {len(all_features)}")

    # Step 2: Convert ArcGIS JSON features to shapely geometries
    coop_geoms = []
    for feat in all_features:
        geom = feat.get("geometry")
        if not geom:
            continue
        try:
            # ArcGIS uses "rings" format — convert to GeoJSON polygon
            rings = geom.get("rings", [])
            if not rings:
                continue
            geojson_geom = {"type": "Polygon", "coordinates": rings}
            shp = shape(geojson_geom)
            if shp.is_valid and not shp.is_empty:
                coop_geoms.append(shp)
        except Exception:
            continue

    log(f"Valid co-op geometries: {len(coop_geoms)}")

    # Step 3: Build spatial index — merge all co-op territories into a single MultiPolygon
    log("Building unified co-op coverage geometry (this may take a minute)...")
    coop_union = unary_union(coop_geoms)
    coop_prepared = prep(coop_union)
    log("Co-op union built successfully")

    # Step 4: Load county boundaries from the project's GeoJSON
    county_geojson_path = Path(__file__).parent.parent / "public" / "data" / "us-counties.json"
    log(f"Loading county boundaries from {county_geojson_path}...")

    with open(county_geojson_path) as f:
        counties_geojson = json.load(f)

    county_features = counties_geojson.get("features", [])
    log(f"Loaded {len(county_features)} county boundaries")

    # Step 5: For each county, compute % area covered by co-op territories
    scores: dict[str, float] = {}
    covered_count = 0
    partial_count = 0

    for feat in county_features:
        props = feat.get("properties", {})
        fips = props.get("FIPS", "")
        if not fips or len(fips) != 5:
            continue

        try:
            county_shape = shape(feat["geometry"])
            if not county_shape.is_valid:
                county_shape = county_shape.buffer(0)

            # Quick check: does any co-op territory touch this county?
            if not coop_prepared.intersects(county_shape):
                scores[fips] = 0.0
                continue

            # Compute intersection area
            intersection = coop_union.intersection(county_shape)
            county_area = county_shape.area
            if county_area <= 0:
                scores[fips] = 0.0
                continue

            ratio = intersection.area / county_area
            score = min(ratio, 1.0)  # Cap at 1.0 (rounding errors)
            scores[fips] = round(score, 4)

            if score >= 0.99:
                covered_count += 1
            elif score > 0:
                partial_count += 1

        except Exception as e:
            log(f"  Warning: geometry error for FIPS {fips}: {e}")
            scores[fips] = 0.0

    zero_coop = sum(1 for s in scores.values() if s == 0.0)
    log(f"Counties scored: {len(scores)}")
    log(f"  Fully covered (≥99%): {covered_count}")
    log(f"  Partially covered: {partial_count}")
    log(f"  Zero coverage: {zero_coop}")

    return scores


def compute_coop_density_legacy(tmpdir: str, fips_lookup: dict) -> dict[str, float]:
    """LEGACY: Compute co-op density as utility count ratio from EIA Form 861.
    Kept for reference — replaced by compute_coop_density_area().
    """
    print("\n=== Computing Co-op Density (EIA-861, LEGACY) ===", flush=True)

    # Download and extract
    zipdata = download(EIA_861_URL, "EIA Form 861 (2024)")
    zf = zipfile.ZipFile(io.BytesIO(zipdata))
    names = zf.namelist()
    log(f"ZIP contains {len(names)} files")

    # Extract needed files
    frame_file = [n for n in names if "Frame" in n and n.endswith(".xlsx")][0]
    territory_file = [n for n in names if "Service_Territory" in n and n.endswith(".xlsx")][0]

    frame_path = os.path.join(tmpdir, "frame.xlsx")
    territory_path = os.path.join(tmpdir, "territory.xlsx")

    with open(frame_path, "wb") as f:
        f.write(zf.read(frame_file))
    with open(territory_path, "wb") as f:
        f.write(zf.read(territory_file))

    # Also extract Reliability file for grid score
    reliability_files = [n for n in names if "Reliability" in n and n.endswith(".xlsx")]
    reliability_path = None
    if reliability_files:
        reliability_path = os.path.join(tmpdir, "reliability.xlsx")
        with open(reliability_path, "wb") as f:
            f.write(zf.read(reliability_files[0]))
        log(f"Extracted reliability file: {reliability_files[0]}")

    zf.close()

    # Build ownership lookup from Frame
    log("Parsing Frame (ownership types)...")
    wb_frame = openpyxl.load_workbook(frame_path, read_only=True)
    ws_frame = wb_frame["Frame"]

    ownership_map: dict[int, str] = {}
    for i, row in enumerate(ws_frame.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header
        util_num = row[1]
        ownership = row[5] if len(row) > 5 else None
        if util_num and ownership:
            ownership_map[util_num] = str(ownership)
    wb_frame.close()
    log(f"Ownership map: {len(ownership_map)} utilities")

    # Parse Service Territory
    log("Parsing Service Territory (county mapping)...")
    wb_terr = openpyxl.load_workbook(territory_path, read_only=True)
    ws_terr = wb_terr["Counties_States"]

    county_utils: dict[str, list[str]] = defaultdict(list)
    matched = 0
    missed = 0

    for i, row in enumerate(ws_terr.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header
        util_num = row[1]
        state = str(row[4]).strip() if row[4] else ""
        county = str(row[5]).strip() if row[5] else ""

        if not state or not county or county == "Not Applicable":
            continue
        if state in SKIP_STATES:
            continue

        ownership = ownership_map.get(util_num, "Unknown")
        fips = resolve_fips(state, county, fips_lookup)

        if fips:
            matched += 1
            county_utils[fips].append(ownership)
        else:
            missed += 1

    wb_terr.close()
    log(f"FIPS match rate: {matched}/{matched + missed} ({100 * matched / max(matched + missed, 1):.1f}%)")

    # Compute co-op density per county
    scores: dict[str, float] = {}
    for fips, ownerships in county_utils.items():
        total = len(ownerships)
        coops = sum(1 for o in ownerships if o == "Cooperative")
        scores[fips] = round(coops / total, 4) if total > 0 else 0.0

    pure_coop = sum(1 for s in scores.values() if s == 1.0)
    zero_coop = sum(1 for s in scores.values() if s == 0.0)
    log(f"Counties scored: {len(scores)} | Pure co-op: {pure_coop} | Zero co-op: {zero_coop}")

    return scores


# ============================================================
# Step 3: Grid Reliability from EIA-861
# ============================================================

def _download_reliability_year(year: int, tmpdir: str) -> str | None:
    """Download EIA 861 ZIP for a single year and extract the Reliability xlsx.

    Returns the path to the extracted reliability file, or None on failure.
    """
    if year == 2024:
        url = EIA_861_CURRENT_URL
    else:
        url = EIA_861_ARCHIVE_URL_TEMPLATE.format(year=year)

    try:
        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
    except Exception as e:
        log(f"  {year}: download failed ({e})")
        return None

    try:
        zf = zipfile.ZipFile(io.BytesIO(resp.content))
        names = zf.namelist()
        reliability_files = [n for n in names if "Reliability" in n and n.endswith(".xlsx")]
        if not reliability_files:
            # Older years might use .xls or different naming
            reliability_files = [n for n in names if "eliability" in n.lower() and (n.endswith(".xlsx") or n.endswith(".xls"))]
        if not reliability_files:
            log(f"  {year}: no reliability file in ZIP ({len(names)} files)")
            zf.close()
            return None
        rel_path = os.path.join(tmpdir, f"reliability_{year}.xlsx")
        with open(rel_path, "wb") as f:
            f.write(zf.read(reliability_files[0]))
        zf.close()
        size_mb = os.path.getsize(rel_path) / 1024 / 1024
        log(f"  {year}: OK ({size_mb:.1f} MB, {reliability_files[0]})")
        return rel_path
    except Exception as e:
        log(f"  {year}: extraction failed ({e})")
        return None


def _parse_reliability_file(rel_path: str) -> dict[int, float]:
    """Parse a single Reliability xlsx and return {utility_number: saidi}."""
    wb_rel = openpyxl.load_workbook(rel_path, read_only=True)

    # Try common sheet names
    sheet_name = None
    for candidate in ["Reliability_States", "Reliability States", "Reliability"]:
        if candidate in wb_rel.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        # Fall back to first sheet
        sheet_name = wb_rel.sheetnames[0]

    ws_rel = wb_rel[sheet_name]

    util_saidi: dict[int, float] = {}
    parsed_count = 0

    for i, row in enumerate(ws_rel.iter_rows(values_only=True)):
        if i < 3:
            continue  # Skip header rows

        vals = list(row)
        if len(vals) < 6:
            continue

        try:
            util_num = int(vals[1])
        except (ValueError, TypeError):
            continue

        parsed_count += 1

        # Try SAIDI sources in priority order:
        # 1. Col 8: IEEE Without Major Event Days (cleanest metric)
        # 2. Col 5: IEEE All Events with MED
        # 3. Col 17: Other Standard All Events
        saidi = None
        for col_idx in [8, 5, 17]:
            raw = vals[col_idx] if col_idx < len(vals) else None
            if raw is None or str(raw).strip() in (".", "", "None"):
                continue
            try:
                val = float(raw)
                if val >= 0:
                    saidi = val
                    break
            except (ValueError, TypeError):
                continue

        if saidi is not None:
            util_saidi[util_num] = saidi

    wb_rel.close()
    return util_saidi


def compute_grid_reliability(tmpdir: str, fips_lookup: dict) -> tuple[dict[str, float], dict[str, dict]]:
    """Compute grid reliability score per county from multi-year EIA-861 Reliability data.

    Downloads reliability data for all available years (2013-2024), averages SAIDI
    across years per county, and returns both scores and metadata about data coverage.

    Returns:
        (scores, metadata) where:
        - scores: {fips: reliability_score}
        - metadata: {fips: {years: [2019, 2020, ...], avg_saidi: 123.4, year_count: 5}}
    """
    print("\n=== Computing Grid Reliability (EIA-861, Multi-Year) ===", flush=True)

    territory_path = os.path.join(tmpdir, "territory.xlsx")
    if not os.path.exists(territory_path):
        log("WARNING: Territory file not found, using fallback estimates")
        return {}, {}

    # Build utility -> counties mapping from 2024 territory data
    log("Building utility-to-county mapping...")
    wb_terr = openpyxl.load_workbook(territory_path, read_only=True)
    ws_terr = wb_terr["Counties_States"]

    util_to_counties: dict[int, set[str]] = defaultdict(set)
    for i, row in enumerate(ws_terr.iter_rows(values_only=True)):
        if i == 0:
            continue
        util_num = row[1]
        state = str(row[4]).strip() if row[4] else ""
        county = str(row[5]).strip() if row[5] else ""
        if not state or not county or state in SKIP_STATES:
            continue
        fips = resolve_fips(state, county, fips_lookup)
        if fips and util_num:
            util_to_counties[util_num].add(fips)
    wb_terr.close()
    log(f"Utilities mapped to counties: {len(util_to_counties)}")

    # Download and parse reliability data for each year
    log(f"Downloading reliability data for {len(EIA_861_RELIABILITY_YEARS)} years...")

    # Per-year SAIDI by utility: {year: {util_num: saidi}}
    yearly_util_saidi: dict[int, dict[int, float]] = {}
    successful_years: list[int] = []

    # 2024 reliability file may already be extracted from the coop density step
    existing_2024 = os.path.join(tmpdir, "reliability.xlsx")
    if os.path.exists(existing_2024):
        log("  2024: using already-extracted file")
        saidi_data = _parse_reliability_file(existing_2024)
        if saidi_data:
            yearly_util_saidi[2024] = saidi_data
            successful_years.append(2024)
            log(f"  2024: {len(saidi_data)} utilities with SAIDI")

    for year in EIA_861_RELIABILITY_YEARS:
        if year in yearly_util_saidi:
            continue  # Already have 2024

        rel_path = _download_reliability_year(year, tmpdir)
        if rel_path is None:
            continue

        saidi_data = _parse_reliability_file(rel_path)
        if saidi_data:
            yearly_util_saidi[year] = saidi_data
            successful_years.append(year)
            log(f"  {year}: {len(saidi_data)} utilities with SAIDI")
        else:
            log(f"  {year}: no SAIDI data parsed")

    successful_years.sort()
    log(f"Successfully loaded {len(successful_years)} years: {successful_years}")

    if not yearly_util_saidi:
        log("WARNING: Could not parse any SAIDI data, using fallback")
        return {}, {}

    # Map SAIDI to counties per year, then average across years
    # county_yearly_saidi: {fips: {year: avg_saidi_for_that_year}}
    county_yearly_saidi: dict[str, dict[int, float]] = defaultdict(dict)

    for year, util_saidi in yearly_util_saidi.items():
        # For each year, average SAIDI across utilities serving each county
        county_year_values: dict[str, list[float]] = defaultdict(list)
        for util_num, saidi in util_saidi.items():
            for fips in util_to_counties.get(util_num, set()):
                county_year_values[fips].append(saidi)

        for fips, values in county_year_values.items():
            county_yearly_saidi[fips][year] = sum(values) / len(values)

    log(f"Counties with any SAIDI data: {len(county_yearly_saidi)}")

    # Compute multi-year average SAIDI per county
    county_avg_saidi: dict[str, float] = {}
    grid_metadata: dict[str, dict] = {}

    for fips, yearly in county_yearly_saidi.items():
        years_with_data = sorted(yearly.keys())
        avg_saidi = sum(yearly.values()) / len(yearly)
        county_avg_saidi[fips] = avg_saidi
        grid_metadata[fips] = {
            "years": years_with_data,
            "year_count": len(years_with_data),
            "min_year": min(years_with_data),
            "max_year": max(years_with_data),
            "avg_saidi": round(avg_saidi, 1),
        }

    if not county_avg_saidi:
        return {}, {}

    # Use percentile-based inverse: accounts for extreme outliers
    fips_list = sorted(county_avg_saidi.keys())
    saidi_vals = [county_avg_saidi[f] for f in fips_list]
    ranks = percentile_rank(saidi_vals)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        # Invert: lowest SAIDI (best reliability) gets highest score
        scores[fips] = round(1.0 - rank, 4)

    # Stats
    saidi_sorted = sorted(county_avg_saidi.values())
    median_saidi = saidi_sorted[len(saidi_sorted) // 2]
    year_counts = [m["year_count"] for m in grid_metadata.values()]
    avg_years = sum(year_counts) / len(year_counts)
    log(f"Counties scored: {len(scores)} | Median SAIDI: {median_saidi:.0f} min/yr")
    log(f"Avg years of data per county: {avg_years:.1f} | Range: {min(year_counts)}-{max(year_counts)}")
    return scores, grid_metadata


# ============================================================
# Step 4: Curtailment Score from EIA-860 + ISO/RTO Data
# ============================================================

def compute_curtailment_score(tmpdir: str, fips_lookup: dict) -> tuple[dict[str, float], dict[int, dict]]:
    """Compute curtailment score from EIA Forms 860 + 923 and ISO/RTO data.

    Methodology (when 923 CF gap data available):
      score = 0.30 * log_norm(installed_MW)
            + 0.15 * pipeline_pressure
            + 0.20 * iso_curtailment_intensity
            + 0.35 * cf_gap_923

    Fallback (when 923 data unavailable):
      score = 0.40 * log_norm(installed_MW)
            + 0.20 * pipeline_pressure
            + 0.40 * iso_curtailment_intensity

    Components:
      - CF gap (35%): Plant-level capacity factor gap from EIA Form 923
        generation data. Most direct measurement of actual curtailment.
      - Installed renewable MW (30%): Log-normalized nameplate capacity from
        EIA Form 860 operable generators (solar PV, solar thermal, wind).
        Captures existing renewable density that drives curtailment.
      - ISO curtailment intensity (20%): Continuous 0-1 score per balancing
        authority from ISO_CURTAILMENT_INTENSITY lookup table, derived from
        2023-2024 ISO/RTO market reports (CAISO, ERCOT, SPP, MISO, PJM).
      - Pipeline pressure (15%): Ratio of proposed-to-existing renewable MW.
        Forward-looking signal for areas where curtailment may worsen.

    Returns: (scores dict, plant_info dict) — plant_info is passed to CF gap.
    """
    print("\n=== Computing Curtailment Proxy (EIA-860) ===", flush=True)

    # Download and extract
    zipdata = download(EIA_860_URL, "EIA Form 860 (2024)")
    zf = zipfile.ZipFile(io.BytesIO(zipdata))
    names = zf.namelist()
    log(f"ZIP contains {len(names)} files")

    plant_file = [n for n in names if "Plant" in n and n.endswith(".xlsx")][0]
    gen_file = [n for n in names if "Generator" in n and "3_1" in n and n.endswith(".xlsx")][0]

    plant_path = os.path.join(tmpdir, "plant.xlsx")
    gen_path = os.path.join(tmpdir, "generator.xlsx")

    with open(plant_path, "wb") as f:
        f.write(zf.read(plant_file))
    with open(gen_path, "wb") as f:
        f.write(zf.read(gen_file))
    zf.close()

    # Parse Plant file for location data
    log("Parsing Plant data (locations)...")
    wb_plant = openpyxl.load_workbook(plant_path, read_only=True)
    ws_plant = wb_plant["Plant"]

    plant_info: dict[int, dict] = {}  # plant_code -> {state, county, ba_code}
    for i, row in enumerate(ws_plant.iter_rows(values_only=True)):
        if i <= 1:
            continue  # Skip title + header rows
        vals = list(row)
        try:
            plant_code = int(vals[2])
        except (ValueError, TypeError):
            continue

        state = str(vals[6]).strip() if vals[6] else ""
        county = str(vals[8]).strip() if vals[8] else ""
        ba_code = str(vals[12]).strip() if vals[12] else ""

        plant_info[plant_code] = {
            "state": state,
            "county": county,
            "ba_code": ba_code,
        }
    wb_plant.close()
    log(f"Plants loaded: {len(plant_info)}")

    # Parse Generator file for renewable capacity
    log("Parsing Generator data (renewable capacity)...")
    wb_gen = openpyxl.load_workbook(gen_path, read_only=True)
    ws_op = wb_gen["Operable"]

    # Find column indices from header row (row 2)
    headers = None
    county_renewable_mw: dict[str, float] = defaultdict(float)
    county_proposed_mw: dict[str, float] = defaultdict(float)
    county_ba_codes: dict[str, set[str]] = defaultdict(set)

    for i, row in enumerate(ws_op.iter_rows(values_only=True)):
        vals = list(row)
        if i == 0:
            continue  # Title row
        if i == 1:
            headers = [str(v).strip() if v else "" for v in vals]
            continue

        if headers is None:
            continue

        # Find needed columns
        tech_idx = next((j for j, h in enumerate(headers) if h == "Technology"), None)
        cap_idx = next((j for j, h in enumerate(headers) if h == "Nameplate Capacity (MW)"), None)
        plant_idx = next((j for j, h in enumerate(headers) if h == "Plant Code"), None)
        status_idx = next((j for j, h in enumerate(headers) if h == "Status"), None)

        if any(idx is None for idx in [tech_idx, cap_idx, plant_idx]):
            continue

        tech = str(vals[tech_idx]) if vals[tech_idx] else ""
        if tech not in VARIABLE_RENEWABLES:
            continue

        status = str(vals[status_idx]) if status_idx and vals[status_idx] else "OP"
        if status not in ("OP", "SB"):
            continue

        try:
            cap = float(vals[cap_idx]) if vals[cap_idx] else 0
        except (ValueError, TypeError):
            cap = 0

        try:
            plant_code = int(vals[plant_idx])
        except (ValueError, TypeError):
            continue

        # Look up county from plant data
        pinfo = plant_info.get(plant_code, {})
        state = pinfo.get("state", "")
        county = pinfo.get("county", "")
        ba_code = pinfo.get("ba_code", "")

        if not state or not county or state in SKIP_STATES:
            continue

        fips = resolve_fips(state, county, fips_lookup)
        if fips:
            county_renewable_mw[fips] += cap
            if ba_code:
                county_ba_codes[fips].add(ba_code)

    wb_gen.close()

    # Also parse Proposed sheet for pipeline pressure
    log("Parsing Proposed generators...")
    wb_gen2 = openpyxl.load_workbook(gen_path, read_only=True)
    ws_prop = wb_gen2["Proposed"]

    prop_headers = None
    for i, row in enumerate(ws_prop.iter_rows(values_only=True)):
        vals = list(row)
        if i == 0:
            continue
        if i == 1:
            prop_headers = [str(v).strip() if v else "" for v in vals]
            continue
        if prop_headers is None:
            continue

        tech_idx = next((j for j, h in enumerate(prop_headers) if h == "Technology"), None)
        cap_idx = next((j for j, h in enumerate(prop_headers) if h == "Nameplate Capacity (MW)"), None)
        plant_idx = next((j for j, h in enumerate(prop_headers) if h == "Plant Code"), None)

        if any(idx is None for idx in [tech_idx, cap_idx, plant_idx]):
            continue

        tech = str(vals[tech_idx]) if vals[tech_idx] else ""
        if tech not in VARIABLE_RENEWABLES:
            continue

        try:
            cap = float(vals[cap_idx]) if vals[cap_idx] else 0
            plant_code = int(vals[plant_idx])
        except (ValueError, TypeError):
            continue

        pinfo = plant_info.get(plant_code, {})
        state = pinfo.get("state", "")
        county = pinfo.get("county", "")
        if state and county and state not in SKIP_STATES:
            fips = resolve_fips(state, county, fips_lookup)
            if fips:
                county_proposed_mw[fips] += cap

    wb_gen2.close()

    log(f"Counties with renewable MW: {len(county_renewable_mw)}")
    log(f"Counties with proposed MW: {len(county_proposed_mw)}")

    # Compute composite curtailment score
    # Components: installed MW (40%), pipeline pressure (20%),
    #             ISO curtailment intensity (40%)
    all_fips = set(county_renewable_mw.keys()) | set(county_proposed_mw.keys())

    if not all_fips:
        return {}, plant_info

    # Normalize renewable MW using log transform (very skewed distribution)
    mw_values = {f: county_renewable_mw.get(f, 0) for f in all_fips}
    log_mw = {f: math.log1p(mw) for f, mw in mw_values.items()}
    max_log_mw = max(log_mw.values()) if log_mw else 1
    norm_mw = {f: v / max_log_mw for f, v in log_mw.items()}

    # Pipeline pressure: proposed / (existing + 1)
    pipeline = {f: county_proposed_mw.get(f, 0) / (county_renewable_mw.get(f, 0) + 1)
                for f in all_fips}
    max_pipeline = max(pipeline.values()) if pipeline else 1
    norm_pipeline = {f: min(v / max(max_pipeline, 1), 1.0) for f, v in pipeline.items()}

    # ISO curtailment intensity from BA codes
    # Use max intensity across all BAs serving generators in this county
    iso_intensity = {}
    for f in all_fips:
        bas = county_ba_codes.get(f, set())
        if bas:
            iso_intensity[f] = max(
                ISO_CURTAILMENT_INTENSITY.get(ba, ISO_CURTAILMENT_DEFAULT)
                for ba in bas
            )
        else:
            iso_intensity[f] = ISO_CURTAILMENT_DEFAULT

    # Compute CF gap from EIA 923 (plant-level generation vs expected)
    cf_gap_scores = compute_cf_gap_923(tmpdir, plant_info, fips_lookup)
    has_923 = len(cf_gap_scores) > 0
    if has_923:
        log(f"CF gap data available for {len(cf_gap_scores)} counties — using 4-component formula")
    else:
        log("CF gap data unavailable — falling back to 3-component formula")

    # Combine components
    scores: dict[str, float] = {}
    for fips in all_fips:
        if has_923:
            # Full formula: 30% MW + 15% pipeline + 20% ISO + 35% CF gap
            score = (
                0.30 * norm_mw.get(fips, 0) +
                0.15 * norm_pipeline.get(fips, 0) +
                0.20 * iso_intensity.get(fips, ISO_CURTAILMENT_DEFAULT) +
                0.35 * cf_gap_scores.get(fips, 0)
            )
        else:
            # Fallback: 40% MW + 20% pipeline + 40% ISO
            score = (
                0.40 * norm_mw.get(fips, 0) +
                0.20 * norm_pipeline.get(fips, 0) +
                0.40 * iso_intensity.get(fips, ISO_CURTAILMENT_DEFAULT)
            )
        scores[fips] = round(min(max(score, 0), 1), 4)

    log(f"Counties with curtailment scores: {len(scores)}")
    return scores, plant_info


# ============================================================
# Step 4b: EIA Form 923 Capacity Factor Gap Analysis
# ============================================================

def compute_cf_gap_923(
    tmpdir: str,
    plant_info: dict[int, dict],
    fips_lookup: dict,
) -> dict[str, float]:
    """Compute plant-level capacity factor gaps from EIA Form 923 generation data.

    Downloads actual monthly generation (Form 923), joins to nameplate capacity
    (Form 860 plant_info), computes actual CF vs expected CF benchmarks, and
    aggregates the capacity-weighted gap to county level.

    A large positive gap means plants in that county are producing significantly
    less than their resource potential — a direct signal of curtailment.

    Returns: dict mapping FIPS -> normalized CF gap score (0-1, higher = more curtailment).
    """
    print("\n--- Computing CF Gap from EIA Form 923 ---", flush=True)

    # Download EIA 923
    try:
        zipdata = download(EIA_923_URL, "EIA Form 923 (2024)")
    except Exception as e:
        log(f"WARNING: Failed to download EIA 923: {e}")
        log("  CF gap analysis will be skipped")
        return {}

    zf = zipfile.ZipFile(io.BytesIO(zipdata))
    names = zf.namelist()
    log(f"ZIP contains {len(names)} files")

    # Find the generation data file — naming varies by year
    gen_file = None
    for name in names:
        lower = name.lower()
        if lower.endswith(".xlsx") and ("generation" in lower or "schedules_2_3_4_5" in lower or "page_1" in lower):
            gen_file = name
            break
    if not gen_file:
        # Fallback: just pick the largest xlsx
        xlsx_files = [n for n in names if n.lower().endswith(".xlsx")]
        if xlsx_files:
            gen_file = xlsx_files[0]

    if not gen_file:
        log("WARNING: No generation spreadsheet found in EIA 923 ZIP")
        return {}

    log(f"Using generation file: {gen_file}")
    gen_path = os.path.join(tmpdir, "eia923_gen.xlsx")
    with open(gen_path, "wb") as f:
        f.write(zf.read(gen_file))
    zf.close()

    # Parse generation data
    log("Parsing EIA 923 generation data...")
    wb = openpyxl.load_workbook(gen_path, read_only=True)

    # Find the right sheet — look for generation/fuel data
    target_sheet = None
    for sn in wb.sheetnames:
        sl = sn.lower()
        if "page 1" in sl or "generation and fuel" in sl or "generation_and_fuel" in sl:
            target_sheet = sn
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]  # Fallback to first sheet

    log(f"Using sheet: {target_sheet}")
    ws = wb[target_sheet]

    # Find header row and column indices
    headers = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip() if v else "" for v in row]
        # Look for a row that has "Plant Id" or "Plant Code" and generation month columns
        has_plant_id = any("plant" in v.lower() and ("id" in v.lower() or "code" in v.lower()) for v in vals)
        has_net_gen = any("net generation" in v.lower() or "netgen" in v.lower() for v in vals)
        if has_plant_id and has_net_gen:
            headers = vals
            header_row_idx = i
            break
        # Also check for month columns like "January", "February" etc
        if has_plant_id and any(v.lower() in ("january", "february", "march") for v in vals):
            headers = vals
            header_row_idx = i
            break
        if i > 10:  # Don't scan too far
            break

    if not headers:
        log("WARNING: Could not find header row in EIA 923 generation sheet")
        wb.close()
        return {}

    log(f"Header row found at index {header_row_idx}")

    # Map column indices
    plant_id_idx = None
    prime_mover_idx = None
    fuel_type_idx = None
    net_gen_month_idxs: list[int] = []

    for j, h in enumerate(headers):
        hl = h.lower()
        if ("plant" in hl and ("id" in hl or "code" in hl)) and plant_id_idx is None:
            plant_id_idx = j
        elif "prime" in hl and "mover" in hl and prime_mover_idx is None:
            prime_mover_idx = j
        elif "fuel" in hl and "type" in hl and "code" in hl and fuel_type_idx is None:
            fuel_type_idx = j
        elif "net" in hl and "gen" in hl:
            net_gen_month_idxs.append(j)
        elif hl in ("january", "february", "march", "april", "may", "june",
                     "july", "august", "september", "october", "november", "december"):
            net_gen_month_idxs.append(j)

    if plant_id_idx is None or not net_gen_month_idxs:
        log(f"WARNING: Missing required columns. plant_id_idx={plant_id_idx}, gen_months={len(net_gen_month_idxs)}")
        wb.close()
        return {}

    log(f"Columns: plant_id={plant_id_idx}, prime_mover={prime_mover_idx}, "
        f"fuel_type={fuel_type_idx}, gen_month_cols={len(net_gen_month_idxs)}")

    # Renewable fuel type codes in EIA 923
    RENEWABLE_FUEL_CODES = {"SUN", "WND"}
    # Renewable prime mover codes
    RENEWABLE_PM_CODES = {"PV", "WT", "CP"}  # PV=solar, WT=wind, CP=concentrated solar

    # Aggregate annual net generation by plant
    # We group by plant_code and track fuel type for expected CF lookup
    plant_annual_gen: dict[int, float] = defaultdict(float)  # plant_code -> total MWh
    plant_fuel_type: dict[int, str] = {}  # plant_code -> "SUN" or "WND"

    data_rows = 0
    renewable_rows = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        vals = list(row)

        # Check fuel type or prime mover for renewable identification
        is_renewable = False
        fuel = ""
        if fuel_type_idx is not None and fuel_type_idx < len(vals):
            fuel = str(vals[fuel_type_idx]).strip().upper() if vals[fuel_type_idx] else ""
            if fuel in RENEWABLE_FUEL_CODES:
                is_renewable = True
        if not is_renewable and prime_mover_idx is not None and prime_mover_idx < len(vals):
            pm = str(vals[prime_mover_idx]).strip().upper() if vals[prime_mover_idx] else ""
            if pm in RENEWABLE_PM_CODES:
                is_renewable = True
                if pm == "PV" or pm == "CP":
                    fuel = "SUN"
                elif pm == "WT":
                    fuel = "WND"

        if not is_renewable:
            continue

        try:
            plant_code = int(vals[plant_id_idx])
        except (ValueError, TypeError):
            continue

        # Sum monthly generation
        annual_gen = 0.0
        valid_months = 0
        for idx in net_gen_month_idxs:
            if idx < len(vals) and vals[idx] is not None:
                try:
                    month_gen = float(vals[idx])
                    annual_gen += month_gen
                    valid_months += 1
                except (ValueError, TypeError):
                    pass

        if valid_months == 0 or annual_gen <= 0:
            continue

        plant_annual_gen[plant_code] += annual_gen
        if plant_code not in plant_fuel_type:
            plant_fuel_type[plant_code] = fuel
        renewable_rows += 1
        data_rows += 1

    wb.close()
    log(f"Parsed {renewable_rows} renewable generation rows for {len(plant_annual_gen)} plants")

    if not plant_annual_gen:
        log("WARNING: No renewable generation data found in EIA 923")
        return {}

    # Now we need nameplate capacity from 860 plant_info
    # The plant_info dict has state/county/ba_code but not capacity.
    # We need to get capacity from the 860 generator data.
    # However, that's already parsed in compute_curtailment_score's operable generators.
    # To avoid re-parsing, we'll compute CF gap per-plant using a simpler approach:
    # re-download isn't needed — we'll parse the already-extracted 860 generator file.
    # But the temp file might not exist. Let's re-extract just the capacity we need.

    # Actually, let's download 860 for just the capacity data
    log("Loading nameplate capacity from EIA 860 for CF calculation...")
    zipdata_860 = download(EIA_860_URL, "EIA Form 860 (for CF gap)")
    zf860 = zipfile.ZipFile(io.BytesIO(zipdata_860))
    gen_860_file = [n for n in zf860.namelist() if "Generator" in n and "3_1" in n and n.endswith(".xlsx")][0]
    gen_860_path = os.path.join(tmpdir, "gen_860_cf.xlsx")
    with open(gen_860_path, "wb") as f:
        f.write(zf860.read(gen_860_file))
    zf860.close()

    # Parse 860 generator data for nameplate capacity by plant
    wb860 = openpyxl.load_workbook(gen_860_path, read_only=True)
    ws860 = wb860["Operable"]

    plant_capacity_mw: dict[int, float] = defaultdict(float)
    plant_op_year: dict[int, int] = {}  # plant_code -> earliest operating year

    headers860 = None
    for i, row in enumerate(ws860.iter_rows(values_only=True)):
        vals = list(row)
        if i == 0:
            continue
        if i == 1:
            headers860 = [str(v).strip() if v else "" for v in vals]
            continue
        if headers860 is None:
            continue

        tech_idx = next((j for j, h in enumerate(headers860) if h == "Technology"), None)
        cap_idx = next((j for j, h in enumerate(headers860) if h == "Nameplate Capacity (MW)"), None)
        plant_idx = next((j for j, h in enumerate(headers860) if h == "Plant Code"), None)
        status_idx = next((j for j, h in enumerate(headers860) if h == "Status"), None)
        op_year_idx = next((j for j, h in enumerate(headers860) if h == "Operating Year"), None)

        if any(idx is None for idx in [tech_idx, cap_idx, plant_idx]):
            continue

        tech = str(vals[tech_idx]) if vals[tech_idx] else ""
        if tech not in VARIABLE_RENEWABLES:
            continue

        status = str(vals[status_idx]) if status_idx is not None and vals[status_idx] else "OP"
        if status not in ("OP", "SB"):
            continue

        try:
            cap = float(vals[cap_idx]) if vals[cap_idx] else 0
            pc = int(vals[plant_idx])
        except (ValueError, TypeError):
            continue

        plant_capacity_mw[pc] += cap

        # Track earliest operating year to filter partial-year plants
        if op_year_idx is not None and vals[op_year_idx]:
            try:
                oy = int(vals[op_year_idx])
                if pc not in plant_op_year or oy < plant_op_year[pc]:
                    plant_op_year[pc] = oy
            except (ValueError, TypeError):
                pass

    wb860.close()
    log(f"Loaded capacity for {len(plant_capacity_mw)} renewable plants from 860")

    # Compute CF gap per plant, aggregate to county
    # CF = annual_gen_mwh / (capacity_mw * 8760)
    # Gap = expected_cf - actual_cf (positive = curtailment)
    hours_in_year = 8760
    county_cf_gap_weighted: dict[str, float] = defaultdict(float)  # fips -> sum(gap * MW)
    county_capacity: dict[str, float] = defaultdict(float)  # fips -> sum(MW)
    plants_matched = 0
    plants_skipped_no_capacity = 0
    plants_skipped_partial_year = 0
    plants_skipped_no_fips = 0

    for plant_code, annual_gen_mwh in plant_annual_gen.items():
        cap_mw = plant_capacity_mw.get(plant_code, 0)
        if cap_mw <= 0:
            plants_skipped_no_capacity += 1
            continue

        # Skip plants that started operating in the data year (partial year)
        op_year = plant_op_year.get(plant_code, 0)
        if op_year >= 2024:
            plants_skipped_partial_year += 1
            continue

        # Get plant location
        pinfo = plant_info.get(plant_code, {})
        state = pinfo.get("state", "")
        county = pinfo.get("county", "")
        if not state or not county or state in SKIP_STATES:
            plants_skipped_no_fips += 1
            continue

        fips = resolve_fips(state, county, fips_lookup)
        if not fips:
            plants_skipped_no_fips += 1
            continue

        # Compute actual capacity factor
        actual_cf = annual_gen_mwh / (cap_mw * hours_in_year)

        # Look up expected CF based on fuel type and state
        fuel = plant_fuel_type.get(plant_code, "")
        if fuel == "SUN":
            expected_cf = EXPECTED_CF_SOLAR.get(state, 0.20)
        elif fuel == "WND":
            expected_cf = EXPECTED_CF_WIND.get(state, 0.30)
        else:
            expected_cf = 0.25  # Generic fallback

        # CF gap: positive means underperforming (curtailed)
        # Clamp to [0, expected_cf] — negative gap means overperforming, not curtailment
        gap = max(0, expected_cf - actual_cf)

        # Weight by plant capacity (larger plants matter more)
        county_cf_gap_weighted[fips] += gap * cap_mw
        county_capacity[fips] += cap_mw
        plants_matched += 1

    log(f"Plants matched: {plants_matched}")
    log(f"Plants skipped: no_capacity={plants_skipped_no_capacity}, "
        f"partial_year={plants_skipped_partial_year}, no_fips={plants_skipped_no_fips}")

    if not county_cf_gap_weighted:
        return {}

    # Compute capacity-weighted average gap per county
    county_avg_gap: dict[str, float] = {}
    for fips in county_cf_gap_weighted:
        if county_capacity[fips] > 0:
            county_avg_gap[fips] = county_cf_gap_weighted[fips] / county_capacity[fips]

    # Normalize via percentile rank
    fips_list = list(county_avg_gap.keys())
    gap_values = [county_avg_gap[f] for f in fips_list]
    ranks = percentile_rank(gap_values)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    log(f"CF gap scores computed for {len(scores)} counties")

    # Report top counties by gap
    top_gap = sorted(county_avg_gap.items(), key=lambda x: x[1], reverse=True)[:10]
    log("Top 10 counties by CF gap (most curtailed):")
    for fips, gap in top_gap:
        log(f"  {fips}: avg_gap={gap:.3f}, capacity={county_capacity[fips]:.0f} MW")

    return scores


# ============================================================
# Step 5: IT Labor from Census CBP
# ============================================================

def compute_labor_score() -> dict[str, float]:
    """Compute IT labor score from Census County Business Patterns."""
    print("\n=== Computing IT Labor Score (Census CBP) ===", flush=True)

    # Fetch CBP data for each NAICS code
    county_emp: dict[str, int] = defaultdict(int)

    for naics in CBP_NAICS_CODES:
        url = f"{CBP_API_BASE}?get=EMP&for=county:*&in=state:*&NAICS2017={naics}&LFO=001&EMPSZES=001"
        log(f"Fetching NAICS {naics}...")

        try:
            resp = requests.get(url, timeout=60)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log(f"WARNING: Failed to fetch NAICS {naics}: {e}")
            continue

        if not data or len(data) < 2:
            log(f"WARNING: No data for NAICS {naics}")
            continue

        for row in data[1:]:
            emp_str = row[0]
            state_fips = row[-2]
            county_fips = row[-1]
            fips = state_fips + county_fips

            try:
                emp = int(emp_str)
            except (ValueError, TypeError):
                continue

            county_emp[fips] += emp

        log(f"  NAICS {naics}: {len(data) - 1} counties")

    log(f"Total counties with IT employment: {len(county_emp)}")

    # Fetch population data
    log("Fetching population estimates...")
    pop_data = download(POP_URL, "Census population estimates")
    pop_text = pop_data.decode("latin-1")
    reader = csv.DictReader(io.StringIO(pop_text))

    county_pop: dict[str, int] = {}
    for row in reader:
        if row["SUMLEV"] != "050":
            continue
        state = row["STATE"].zfill(2)
        county = row["COUNTY"].zfill(3)
        fips = state + county
        try:
            pop = int(row["POPESTIMATE2023"])
        except (ValueError, KeyError):
            try:
                pop = int(row["POPESTIMATE2024"])
            except (ValueError, KeyError):
                continue
        county_pop[fips] = pop

    log(f"Counties with population: {len(county_pop)}")

    # Compute per-capita IT density and normalize via percentile rank
    all_fips = set(county_pop.keys())
    densities: dict[str, float] = {}

    for fips in all_fips:
        emp = county_emp.get(fips, 0)
        pop = county_pop.get(fips, 1)
        if pop > 0:
            densities[fips] = emp / pop * 10000  # per 10K residents
        else:
            densities[fips] = 0.0

    # Percentile rank normalization
    fips_list = sorted(densities.keys())
    density_vals = [densities[f] for f in fips_list]
    ranks = percentile_rank(density_vals)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    positive = sum(1 for f in scores if county_emp.get(f, 0) > 0)
    log(f"Counties scored: {len(scores)} | With IT employment: {positive}")
    return scores


def blend_labor_with_neighbors(
    labor_scores: dict[str, float],
    self_weight: float = 0.75,
) -> dict[str, float]:
    """Blend labor scores with neighboring county averages.

    Labor markets don't stop at county lines — a site in a rural county
    adjacent to a metro area benefits from the metro's workforce. This
    function computes neighbor-blended scores using county boundary
    adjacency from the project's GeoJSON.

    Formula: blended = self_weight × own_score + (1 - self_weight) × avg(neighbor_scores)

    Uses shapely STRtree for efficient spatial neighbor lookup.
    """
    print("\n--- Blending labor scores with neighboring counties ---", flush=True)

    county_geojson_path = Path(__file__).parent.parent / "public" / "data" / "us-counties.json"
    with open(county_geojson_path) as f:
        counties_geojson = json.load(f)

    features = counties_geojson.get("features", [])

    # Build county shapes and spatial index
    shapes_by_fips: dict[str, object] = {}
    fips_list: list[str] = []
    geom_list: list[object] = []

    for feat in features:
        props = feat.get("properties", {})
        fips = props.get("FIPS", "")
        if not fips or len(fips) != 5:
            continue
        try:
            s = shape(feat["geometry"])
            if not s.is_valid:
                s = s.buffer(0)
            shapes_by_fips[fips] = s
            fips_list.append(fips)
            geom_list.append(s)
        except Exception:
            continue

    log(f"Built shapes for {len(fips_list)} counties")

    # Build STRtree spatial index for fast neighbor queries
    tree = STRtree(geom_list)

    # For each county, find neighbors (counties whose boundaries touch/intersect)
    neighbor_weight = 1.0 - self_weight
    blended: dict[str, float] = {}
    blend_count = 0

    for i, fips in enumerate(fips_list):
        own_score = labor_scores.get(fips, 0.0)
        county_geom = geom_list[i]

        # Query tree for geometries that intersect a small buffer of the county boundary
        # Using the boundary (not the polygon) to find touching/adjacent counties
        boundary = county_geom.boundary.buffer(0.001)  # ~100m buffer for numerical stability
        candidate_indices = tree.query(boundary)

        neighbor_scores: list[float] = []
        for idx in candidate_indices:
            neighbor_fips = fips_list[idx]
            if neighbor_fips == fips:
                continue  # Skip self
            # Verify actual intersection (tree query returns bbox candidates)
            if county_geom.touches(geom_list[idx]) or county_geom.boundary.intersects(geom_list[idx].boundary):
                ns = labor_scores.get(neighbor_fips, 0.0)
                neighbor_scores.append(ns)

        if neighbor_scores:
            avg_neighbor = sum(neighbor_scores) / len(neighbor_scores)
            blended[fips] = round(self_weight * own_score + neighbor_weight * avg_neighbor, 4)
            if blended[fips] != round(own_score, 4):
                blend_count += 1
        else:
            blended[fips] = round(own_score, 4)

    # Include any counties from labor_scores that weren't in the GeoJSON
    for fips, score in labor_scores.items():
        if fips not in blended:
            blended[fips] = round(score, 4)

    log(f"Blended {blend_count} counties with neighbor influence (self={self_weight}, neighbor={neighbor_weight})")
    return blended


# ============================================================
# Step 6: Fiber / Broadband from Census ACS
# ============================================================

def compute_fiber_fcc_bdc(
    ntia_scores: dict[str, float] | None = None,
) -> tuple[dict[str, float], dict[str, dict]]:
    """Compute fiber availability from FCC BDC county-level data (Dec 2024).

    Primary fiber source: ISP-reported fiber-to-the-premises availability
    at the location level, aggregated to county via ArcGIS Living Atlas.

    When ntia_scores are provided, the composite formula changes from
    80/20 (BDC rate / providers) to 60/15/25 (BDC rate / providers / NTIA).

    Returns:
        (scores, metadata) where:
        - scores: {fips: fiber_score}
        - metadata: {fips: {pct_fiber: 0.53, total_bsls: 12345, fiber_bsls: 6543, providers: 3}}
    """
    print("\n=== Computing Fiber Score (FCC BDC, Dec 2024) ===", flush=True)

    fields = "GEOID,TotalBSLs,ServedBSLsFiber,UnderservedBSLsFiber,UnservedBSLsFiber,UniqueProvidersFiber"
    all_features: list[dict] = []

    # ArcGIS paginates at 2000 records max; total is ~3234
    offset = 0
    batch_size = 2000
    while True:
        url = (
            f"{FCC_BDC_ARCGIS_URL}?where=1%3D1&outFields={fields}"
            f"&resultRecordCount={batch_size}&resultOffset={offset}"
            f"&f=json&returnGeometry=false"
        )
        log(f"Fetching BDC county data (offset {offset})...")
        try:
            resp = requests.get(url, timeout=120)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log(f"WARNING: FCC BDC fetch failed: {e}")
            break

        features = data.get("features", [])
        if not features:
            break
        all_features.extend(features)
        if len(features) < batch_size:
            break
        offset += batch_size

    log(f"FCC BDC counties fetched: {len(all_features)}")

    if not all_features:
        log("WARNING: No FCC BDC data, will fall back to ACS")
        return {}, {}

    # Compute fiber availability rate per county
    # "Fiber available" = locations where fiber is served OR underserved (i.e., fiber exists but < 100/20)
    # This captures actual fiber infrastructure presence, not just high-speed fiber
    fiber_rates: dict[str, float] = {}
    fiber_metadata: dict[str, dict] = {}

    for feat in all_features:
        a = feat["attributes"]
        geoid = str(a.get("GEOID", "")).zfill(5)
        total_bsls = a.get("TotalBSLs") or 0
        served_fiber = a.get("ServedBSLsFiber") or 0
        underserved_fiber = a.get("UnderservedBSLsFiber") or 0
        providers = a.get("UniqueProvidersFiber") or 0

        if total_bsls <= 0:
            continue

        # Fiber-available locations = served + underserved (fiber exists at location)
        fiber_bsls = served_fiber + underserved_fiber
        pct = fiber_bsls / total_bsls

        fiber_rates[geoid] = pct
        fiber_metadata[geoid] = {
            "pct_fiber": round(pct, 4),
            "total_bsls": total_bsls,
            "fiber_bsls": fiber_bsls,
            "fiber_providers": providers,
        }

    log(f"Counties with FCC BDC fiber data: {len(fiber_rates)}")

    if not fiber_rates:
        return {}, {}

    # Composite score weights depend on whether NTIA middle-mile data is available
    # Without NTIA: 80% fiber availability rate + 20% provider competition
    # With NTIA:    60% fiber availability rate + 15% provider competition + 25% NTIA middle mile
    has_ntia = ntia_scores and len(ntia_scores) > 0
    if has_ntia:
        log("Blending NTIA Middle Mile data into fiber composite (60/15/25 weights)")
    else:
        log("No NTIA data — using standard fiber composite (80/20 weights)")

    fips_list = sorted(fiber_rates.keys())
    raw_scores: list[float] = []
    for fips in fips_list:
        rate = fiber_rates[fips]
        providers = fiber_metadata[fips]["fiber_providers"]
        # Normalize providers: diminishing returns after ~5
        provider_score = min(providers / 5.0, 1.0)
        if has_ntia:
            ntia = ntia_scores.get(fips, 0.0)
            composite = 0.60 * rate + 0.15 * provider_score + 0.25 * ntia
        else:
            composite = 0.80 * rate + 0.20 * provider_score
        raw_scores.append(composite)

    # Percentile rank normalization
    ranks = percentile_rank(raw_scores)
    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    # Stats
    rates_sorted = sorted(fiber_rates.values())
    median_rate = rates_sorted[len(rates_sorted) // 2]
    zero_fiber = sum(1 for r in fiber_rates.values() if r == 0)
    high_fiber = sum(1 for r in fiber_rates.values() if r > 0.8)
    log(f"Median fiber availability: {median_rate:.1%}")
    log(f"Zero fiber: {zero_fiber} counties | >80% fiber: {high_fiber} counties")

    return scores, fiber_metadata


def compute_fiber_acs_fallback() -> dict[str, float]:
    """Compute fiber proxy from Census ACS broadband subscriptions (fallback only)."""
    print("\n=== Computing Fiber Fallback (Census ACS Broadband) ===", flush=True)

    variables = "B28002_001E,B28002_004E,B28002_007E"
    url = f"{ACS_BROADBAND_URL}?get=NAME,{variables}&for=county:*&in=state:*"

    log("Fetching ACS broadband subscription data...")
    try:
        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        log(f"WARNING: Failed to fetch ACS broadband data: {e}")
        return {}

    if not data or len(data) < 2:
        return {}

    broadband_rates: dict[str, float] = {}
    for row in data[1:]:
        try:
            total_hh = int(row[1]) if row[1] else 0
            cable_fiber = int(row[2]) if row[2] else 0
            state_fips = row[-2]
            county_fips = row[-1]
            fips = state_fips + county_fips
            if total_hh > 0:
                broadband_rates[fips] = cable_fiber / total_hh
        except (ValueError, TypeError, IndexError):
            continue

    if not broadband_rates:
        return {}

    fips_list = sorted(broadband_rates.keys())
    rate_vals = [broadband_rates[f] for f in fips_list]
    ranks = percentile_rank(rate_vals)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    log(f"ACS fallback counties: {len(scores)}")
    return scores


def compute_fiber_score(
    ntia_scores: dict[str, float] | None = None,
) -> tuple[dict[str, float], dict[str, dict], dict[str, str]]:
    """Compute fiber score using FCC BDC as primary, ACS as fallback.

    When ntia_scores are provided, they are blended into the BDC composite
    at 25% weight (reducing BDC rate from 80% to 60% and providers from 20% to 15%).

    Returns:
        (scores, metadata, source_map) where:
        - scores: {fips: fiber_score}
        - metadata: {fips: {...}} (only for BDC counties)
        - source_map: {fips: "FCC BDC Dec 2024..." | "Census ACS 2023 (fallback)"}
    """
    # Primary: FCC BDC (with optional NTIA blending)
    bdc_scores, bdc_metadata = compute_fiber_fcc_bdc(ntia_scores=ntia_scores)

    # Fallback: Census ACS
    acs_scores = compute_fiber_acs_fallback()

    # Merge: BDC takes priority
    scores: dict[str, float] = {}
    source_map: dict[str, str] = {}

    all_fips = set(bdc_scores.keys()) | set(acs_scores.keys())
    bdc_used = 0
    acs_used = 0

    bdc_source_label = (
        "FCC BDC Dec 2024 + NTIA Middle Mile grants (USAspending.gov)"
        if ntia_scores
        else "FCC BDC Dec 2024 (fiber availability at BSLs via ArcGIS Living Atlas)"
    )

    for fips in all_fips:
        if fips in bdc_scores:
            scores[fips] = bdc_scores[fips]
            source_map[fips] = bdc_source_label
            bdc_used += 1
        elif fips in acs_scores:
            scores[fips] = acs_scores[fips]
            source_map[fips] = "Census ACS 2023 (broadband subscription fallback)"
            acs_used += 1

    log(f"\nFiber scoring summary: {bdc_used} BDC + {acs_used} ACS fallback = {len(scores)} total")
    return scores, bdc_metadata, source_map


# ============================================================
# Step 6a-ii: NTIA Enabling Middle Mile Broadband Infrastructure
# ============================================================

def _parse_counties_from_description(
    description: str, pop_state: str | None, fips_lookup: dict
) -> list[tuple[str, str]]:
    """Parse county names and states from an NTIA grant description.

    Descriptions are ALL-CAPS and use varied formats:
      - "BARRON, POLK, AND WASHBURN COUNTIES IN WISCONSIN"
      - "LOVING COUNTY, TX"
      - "BERRIEN COUNTY"
      - "CUMBERLAND AND SALEM COUNTIES"

    Returns list of (state_abbr, county_name) pairs.
    """
    text = description.upper()
    results: list[tuple[str, str]] = []

    # Pattern 1: "[county], [county], AND [county] COUNTIES IN [state]"
    # e.g., "BARRON, POLK, BURNETT, AND WASHBURN COUNTIES IN WISCONSIN"
    pat1 = re.findall(
        r'((?:[A-Z][A-Z .\']+(?:,\s*)?)+(?:,?\s*AND\s+[A-Z][A-Z .\']+))\s+COUNT(?:Y|IES)\s+IN\s+'
        r'((?:NEW\s+)?(?:NORTH\s+)?(?:SOUTH\s+)?(?:WEST\s+)?(?:RHODE\s+)?[A-Z]+)',
        text,
    )
    for county_group, state_name in pat1:
        state_name = state_name.strip()
        abbr = STATE_NAME_TO_ABBR.get(state_name)
        if not abbr:
            continue
        # Split "BARRON, POLK, BURNETT, AND WASHBURN"
        names = re.split(r',\s*|\s+AND\s+', county_group)
        for name in names:
            name = name.strip().rstrip(',')
            if name and len(name) > 1:
                results.append((abbr, name.title()))

    # Pattern 2: "[county] COUNTY, [state_abbr]"
    # e.g., "LOVING COUNTY, TX"
    pat2 = re.findall(r'([A-Z][A-Z .\']+?)\s+COUNTY,?\s+([A-Z]{2})\b', text)
    for county_name, state_abbr in pat2:
        county_name = county_name.strip()
        if state_abbr in STATE_NAME_TO_ABBR.values() and len(county_name) > 1:
            results.append((state_abbr, county_name.title()))

    # Pattern 3: "[county] AND [county] COUNTIES" (no state — use pop_state)
    if pop_state:
        pat3 = re.findall(
            r'([A-Z][A-Z .\']+?)\s+AND\s+([A-Z][A-Z .\']+?)\s+COUNT(?:Y|IES)',
            text,
        )
        for c1, c2 in pat3:
            c1, c2 = c1.strip(), c2.strip()
            if len(c1) > 1:
                results.append((pop_state, c1.title()))
            if len(c2) > 1:
                results.append((pop_state, c2.title()))

    # Pattern 4: Standalone "[county] COUNTY" (singular, use pop_state)
    if pop_state:
        pat4 = re.findall(r'(?<![,\w])\s([A-Z][A-Z .\']{2,}?)\s+COUNTY(?!\s*,?\s*[A-Z]{2}\b)', text)
        for county_name in pat4:
            county_name = county_name.strip()
            # Skip false positives (common words that precede COUNTY)
            if county_name in ("THE", "EACH", "EVERY", "ONE", "THIS", "THAT", "PER", "UNSERVED"):
                continue
            if len(county_name) > 1:
                results.append((pop_state, county_name.title()))

    # Deduplicate
    seen: set[tuple[str, str]] = set()
    deduped: list[tuple[str, str]] = []
    for pair in results:
        key = (pair[0], pair[1].lower())
        if key not in seen:
            seen.add(key)
            deduped.append(pair)

    return deduped


def compute_ntia_middle_mile(
    fips_lookup: dict,
) -> tuple[dict[str, float], dict[str, dict]]:
    """Fetch NTIA Enabling Middle Mile grants and compute per-county scores.

    Uses USAspending.gov API to fetch all CFDA 11.033 grants, then maps
    each grant to the counties it serves using Place of Performance data
    and description parsing.

    Returns:
        (scores, metadata) where:
        - scores: {fips: ntia_score} (0-1, log-normalized grant amount)
        - metadata: {fips: {grant_count, total_amount, recipients: [...]}}
    """
    print("\n=== Computing NTIA Middle Mile Score (USAspending.gov) ===", flush=True)

    # Step 1: Fetch all grants from search API
    log("Fetching NTIA Middle Mile grants (CFDA 11.033)...")
    payload = {
        "filters": {"program_numbers": [NTIA_MIDDLE_MILE_CFDA]},
        "fields": [
            "Award ID", "Recipient Name", "Award Amount",
            "Place of Performance State Code",
            "generated_internal_id",
        ],
        "limit": 100,
        "page": 1,
        "sort": "Award Amount",
        "order": "desc",
    }

    try:
        resp = requests.post(USASPENDING_SEARCH_URL, json=payload, timeout=60)
        resp.raise_for_status()
        search_data = resp.json()
    except Exception as e:
        log(f"WARNING: USAspending search failed: {e}")
        return {}, {}

    results = search_data.get("results", [])
    log(f"Found {len(results)} NTIA Middle Mile grants")

    if not results:
        return {}, {}

    # Step 2: Fetch individual award details for geographic data + descriptions
    grants: list[dict] = []
    for i, award in enumerate(results):
        amount = award.get("Award Amount") or 0
        if amount <= 0:
            continue

        internal_id = award.get("generated_internal_id", "")
        recipient = award.get("Recipient Name", "Unknown")
        pop_state = award.get("Place of Performance State Code")

        # Fetch award detail for description and county-level geography
        detail = None
        if internal_id:
            try:
                detail_url = f"{USASPENDING_AWARD_URL}{internal_id}/"
                detail_resp = requests.get(detail_url, timeout=30)
                detail_resp.raise_for_status()
                detail = detail_resp.json()
            except Exception:
                pass  # Fall back to search-level data

        description = ""
        pop_county_code = None
        pop_state_code = None

        if detail:
            description = detail.get("description", "") or ""
            pop = detail.get("place_of_performance") or {}
            pop_county_code = pop.get("county_code")
            pop_state_code = pop.get("state_code") or pop_state

            # Also check recipient location for state
            recipient_loc = (detail.get("recipient") or {}).get("location") or {}
            if not pop_state_code:
                pop_state_code = recipient_loc.get("state_code")
        else:
            pop_state_code = pop_state

        grants.append({
            "award_id": award.get("Award ID", ""),
            "recipient": recipient,
            "amount": amount,
            "description": description,
            "pop_state": pop_state_code,
            "pop_county_code": pop_county_code,
        })

        if (i + 1) % 10 == 0:
            log(f"  Fetched details for {i + 1}/{len(results)} grants...")

    log(f"Grants with non-zero amounts: {len(grants)}")

    # Step 3: Map grants to counties
    # For each grant, resolve: (a) Place of Performance county, (b) counties parsed from description
    county_grants: dict[str, list[dict]] = defaultdict(list)  # fips -> [grant_info, ...]

    for grant in grants:
        matched_fips: set[str] = set()

        # (a) Place of Performance county (single county — HQ or primary site)
        if grant["pop_county_code"] and grant["pop_state"]:
            # Build 5-digit FIPS from state abbreviation + 3-digit county code
            state_abbr = grant["pop_state"]
            # We need state FIPS from abbreviation — scan fips_lookup
            state_fips = None
            for (st, _), fips_code in fips_lookup.items():
                if st == state_abbr:
                    state_fips = fips_code[:2]
                    break
            if state_fips:
                pop_fips = state_fips + grant["pop_county_code"].zfill(3)
                matched_fips.add(pop_fips)

        # (b) Parse county names from description
        if grant["description"]:
            parsed_counties = _parse_counties_from_description(
                grant["description"], grant["pop_state"], fips_lookup,
            )
            for state_abbr, county_name in parsed_counties:
                fips = resolve_fips(state_abbr, county_name, fips_lookup)
                if fips:
                    matched_fips.add(fips)

        # If no counties matched, try the pop_state as a state-wide signal
        # (spreads to no specific county — skip these)
        if not matched_fips:
            log(f"  WARNING: No counties matched for {grant['recipient']} "
                f"(${grant['amount']:,.0f})")
            continue

        # Classify recipient type
        recipient_lower = grant["recipient"].lower()
        if any(kw in recipient_lower for kw in ("cooperative", "co-op", "coop")):
            rtype = "cooperative"
        elif any(kw in recipient_lower for kw in ("utility", "power", "electric", "energy")):
            rtype = "utility"
        elif any(kw in recipient_lower for kw in ("tel", "telecom", "communications", "broadband", "fiber")):
            rtype = "telco"
        else:
            rtype = "other"

        grant_info = {
            "recipient": grant["recipient"],
            "amount": grant["amount"],
            "recipient_type": rtype,
            "award_id": grant["award_id"],
        }

        # Assign grant to all matched counties
        for fips in matched_fips:
            county_grants[fips].append(grant_info)

    log(f"Counties with NTIA Middle Mile grants: {len(county_grants)}")

    if not county_grants:
        return {}, {}

    # Step 4: Compute per-county scores
    # Score = log-normalized total grant amount (similar to queue_pressure)
    county_amounts: dict[str, float] = {}
    ntia_metadata: dict[str, dict] = {}

    for fips, grant_list in county_grants.items():
        total_amount = sum(g["amount"] for g in grant_list)
        recipients = list({g["recipient"] for g in grant_list})
        recipient_types = list({g["recipient_type"] for g in grant_list})

        county_amounts[fips] = total_amount
        ntia_metadata[fips] = {
            "grant_count": len(grant_list),
            "total_amount": round(total_amount, 2),
            "recipients": recipients,
            "recipient_types": recipient_types,
        }

    # Log-normalize and percentile-rank
    fips_list = sorted(county_amounts.keys())
    log_amounts = [math.log1p(county_amounts[f]) for f in fips_list]
    ranks = percentile_rank(log_amounts)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    # Stats
    total_grant_dollars = sum(county_amounts.values())
    coop_counties = sum(
        1 for m in ntia_metadata.values()
        if "cooperative" in m["recipient_types"]
    )
    log(f"Total grant dollars mapped: ${total_grant_dollars:,.0f}")
    log(f"Counties with cooperative recipients: {coop_counties}")

    # Spot-check Dairyland
    for fips, meta in ntia_metadata.items():
        for r in meta["recipients"]:
            if "DAIRYLAND" in r.upper():
                log(f"  Dairyland Power mapped to FIPS {fips} "
                    f"(${meta['total_amount']:,.0f}, score={scores.get(fips, 0):.4f})")

    return scores, ntia_metadata


# ============================================================
# Step 6b: Permitting Score (state policy + risk + county adjustments)
# ============================================================

# State-level data center policy scores (tax incentives, exemptions)
# Sources: NCSL Policy Snapshot, SDI Alliance, H5 Data Centers, NAIOP (2025-2026)
_STATE_DC_POLICY = {
    "AL": 0.85, "AZ": 0.70, "CT": 0.80, "FL": 0.85, "GA": 0.85,
    "IA": 0.80, "IN": 0.85, "KS": 0.65, "KY": 0.65, "LA": 0.55,
    "MN": 0.65, "MS": 0.70, "MO": 0.75, "NB": 0.75, "NE": 0.75,
    "NV": 0.80, "NY": 0.60, "NC": 0.80, "ND": 0.65, "OH": 0.75,
    "OK": 0.65, "SC": 0.80, "TN": 0.75, "TX": 0.80, "VA": 0.80,
    "WA": 0.70, "WV": 0.80, "WY": 0.75,
    # Limited / no DC-specific incentives
    "CO": 0.45, "WI": 0.40, "OR": 0.55, "DE": 0.55, "MT": 0.50,
    "NH": 0.50, "SD": 0.45, "UT": 0.45, "PA": 0.35, "IL": 0.30,
    "ID": 0.30, "AK": 0.40, "AR": 0.25, "CA": 0.20, "HI": 0.15,
    "ME": 0.25, "MD": 0.30, "MA": 0.25, "MI": 0.30, "NJ": 0.25,
    "NM": 0.30, "RI": 0.20, "VT": 0.20,
}

# Regulatory environment (lower burden = higher score)
_STATE_REGULATORY = {
    "TX": 0.90, "FL": 0.85, "GA": 0.85, "TN": 0.85, "AL": 0.80,
    "SC": 0.80, "NC": 0.80, "IN": 0.85, "OH": 0.75, "MO": 0.80,
    "MS": 0.80, "OK": 0.80, "AR": 0.75, "LA": 0.70, "KY": 0.75,
    "WV": 0.75, "KS": 0.80, "NE": 0.80, "SD": 0.85, "ND": 0.80,
    "WY": 0.85, "MT": 0.75, "ID": 0.80, "UT": 0.80, "AZ": 0.75,
    "NV": 0.80, "IA": 0.80,
    "VA": 0.65, "PA": 0.55, "MI": 0.60, "WI": 0.65, "MN": 0.60,
    "CO": 0.60, "NM": 0.60, "WA": 0.55, "OR": 0.50, "DE": 0.65,
    "MD": 0.55, "NH": 0.70, "AK": 0.65,
    "NY": 0.45, "NJ": 0.40, "CT": 0.50, "MA": 0.40, "IL": 0.45,
    "CA": 0.25, "HI": 0.30, "VT": 0.45, "ME": 0.55, "RI": 0.40,
}

# Risk factors inverted (1.0 = low risk). Sources: Data Center Watch 2025
_STATE_RISK = {
    "TX": 0.85, "GA": 0.80, "TN": 0.85, "AL": 0.85, "SC": 0.80,
    "NC": 0.75, "FL": 0.80, "MS": 0.85, "OK": 0.85, "AR": 0.85,
    "LA": 0.80, "WV": 0.75, "NE": 0.85, "SD": 0.90, "ND": 0.85,
    "WY": 0.90, "MT": 0.85, "ID": 0.80, "UT": 0.80, "NV": 0.80,
    "KS": 0.75, "NM": 0.80, "AK": 0.85, "DE": 0.80, "NH": 0.80,
    "IA": 0.80, "KY": 0.70, "OH": 0.70, "PA": 0.70, "MI": 0.65,
    "WI": 0.70, "MN": 0.60, "CO": 0.70, "WA": 0.65, "MD": 0.70,
    "CT": 0.70, "IN": 0.65, "MO": 0.60, "NY": 0.60,
    "VA": 0.40, "AZ": 0.55, "OR": 0.50, "CA": 0.50, "IL": 0.55,
    "NJ": 0.60, "MA": 0.55, "VT": 0.65, "ME": 0.70, "RI": 0.60,
    "HI": 0.50,
}

# Zoning favorability
_STATE_ZONING = {
    "TX": 0.90, "WY": 0.90, "MT": 0.85, "SD": 0.85, "ND": 0.85,
    "NE": 0.85, "KS": 0.85, "OK": 0.85, "IA": 0.85, "ID": 0.80,
    "NV": 0.80, "UT": 0.80, "AR": 0.80, "MS": 0.80, "AL": 0.80,
    "NM": 0.75, "AK": 0.70, "WI": 0.75, "MN": 0.75, "CO": 0.75,
    "GA": 0.75, "SC": 0.75, "NC": 0.75, "TN": 0.75, "IN": 0.75,
    "OH": 0.70, "MO": 0.70, "KY": 0.75, "WV": 0.75, "LA": 0.70,
    "FL": 0.70, "VA": 0.65, "PA": 0.65, "MI": 0.70, "WA": 0.65,
    "OR": 0.60, "AZ": 0.70, "ME": 0.70, "NH": 0.65, "VT": 0.65,
    "DE": 0.60, "NY": 0.50, "NJ": 0.45, "CT": 0.55, "MA": 0.50,
    "MD": 0.55, "IL": 0.60, "CA": 0.45, "HI": 0.35, "RI": 0.45,
}

# County-level adjustments (blocked projects, moratoriums, local factors)
_COUNTY_ADJUSTMENTS = {
    "51153": -0.15, "51047": -0.10, "51099": -0.10, "51061": -0.10,
    "51107": -0.08, "51059": -0.08, "51087": -0.05,  # Virginia opposition
    "29037": -0.15,  # Cass County MO (Peculiar moratorium)
    "18127": -0.10,  # Porter County IN (Chesterton blocked)
    "04013": -0.05,  # Maricopa AZ (mixed)
    "41027": -0.10,  # Hood River OR (Cascade Locks)
    "55033": 0.05, "55123": 0.05,  # Dunn/Vernon WI (rural co-op friendly)
    "08123": 0.03,  # Weld CO (industrial-friendly)
}


def compute_permitting_score(state: str, fips: str) -> float:
    """Compute permitting score from state policy, regulation, risk, and zoning."""
    policy = _STATE_DC_POLICY.get(state, 0.35)
    regulatory = _STATE_REGULATORY.get(state, 0.50)
    risk = _STATE_RISK.get(state, 0.70)
    zoning = _STATE_ZONING.get(state, 0.60)

    base = 0.40 * policy + 0.25 * regulatory + 0.20 * risk + 0.15 * zoning
    adj = _COUNTY_ADJUSTMENTS.get(fips, 0.0)
    return round(max(0.05, min(0.95, base + adj)), 4)


# ============================================================
# Step 7: Negative LMP Frequency (gridstatus)
# ============================================================

def compute_negative_lmp_frequency(
    plant_info: dict[int, dict],
    fips_lookup: dict,
) -> dict[str, float]:
    """Compute negative LMP frequency per county using gridstatus library.

    Queries zone-level day-ahead hourly LMP from each supported ISO for a
    recent 90-day window. Computes % of hours with negative prices per zone,
    then maps zones to counties via BA codes from the 860 plant data.

    Returns: dict mapping FIPS -> negative LMP frequency score (0-1, higher = more negative prices).
    Falls back to empty dict if gridstatus is not available or queries fail.
    """
    print("\n=== Computing Negative LMP Frequency (gridstatus) ===", flush=True)

    try:
        import gridstatus
        import pandas as pd
    except ImportError:
        log("WARNING: gridstatus or pandas not installed, skipping negative LMP analysis")
        log("  Install with: uv add gridstatus pandas")
        return {}

    # Build BA code -> set of FIPS mapping from plant_info
    ba_to_fips: dict[str, set[str]] = defaultdict(set)
    for plant_code, pinfo in plant_info.items():
        ba = pinfo.get("ba_code", "")
        state = pinfo.get("state", "")
        county = pinfo.get("county", "")
        if ba and state and county and state not in SKIP_STATES:
            fips = resolve_fips(state, county, fips_lookup)
            if fips:
                ba_to_fips[ba].add(fips)
    log(f"BA codes mapped to counties: {len(ba_to_fips)}")

    # Query each ISO for zone-level day-ahead hourly LMP
    # Use 90-day window ending yesterday
    end_date = pd.Timestamp.now().normalize() - pd.Timedelta(days=1)
    start_date = end_date - pd.Timedelta(days=90)
    log(f"Query window: {start_date.date()} to {end_date.date()}")

    # ISO class -> BA code prefix mapping
    iso_configs: list[tuple[str, object, str, dict]] = []
    try:
        iso_configs = [
            ("CAISO", gridstatus.CAISO(), "DAY_AHEAD_HOURLY", {"location_type": "ZONE"}),
            ("PJM", gridstatus.PJM(), "REAL_TIME_HOURLY", {"location_type": "ZONE"}),
            ("MISO", gridstatus.MISO(), "DAY_AHEAD_HOURLY", {}),
            ("SPP", gridstatus.SPP(), "DAY_AHEAD_HOURLY", {}),
            ("NYISO", gridstatus.NYISO(), "DAY_AHEAD_5_MIN", {"location_type": "ZONE"}),
            ("ISONE", gridstatus.ISONE(), "DAY_AHEAD_HOURLY", {"location_type": "ZONE"}),
        ]
    except Exception as e:
        log(f"WARNING: Failed to initialize gridstatus ISOs: {e}")
        return {}

    # BA code -> ISO name mapping for county attribution
    ba_to_iso = {
        "CISO": "CAISO", "PJM": "PJM", "MISO": "MISO",
        "SWPP": "SPP", "NYIS": "NYISO", "ISNE": "ISONE",
        "ERCO": "ERCOT",
    }

    # Collect negative price percentage per ISO
    iso_neg_pct: dict[str, float] = {}

    for iso_name, iso_obj, market, kwargs in iso_configs:
        try:
            log(f"  Querying {iso_name} zone-level LMP ({market})...")
            df = iso_obj.get_lmp(
                start=start_date,
                end=end_date,
                market=market,
                **kwargs,
            )
            if df is not None and len(df) > 0:
                # Count negative price hours
                lmp_col = "LMP" if "LMP" in df.columns else df.columns[df.columns.str.contains("LMP", case=False)].tolist()
                if isinstance(lmp_col, list) and lmp_col:
                    lmp_col = lmp_col[0]
                elif isinstance(lmp_col, str):
                    pass
                else:
                    log(f"    No LMP column found in {iso_name} data")
                    continue

                total_rows = len(df)
                neg_rows = (df[lmp_col] < 0).sum()
                neg_pct = neg_rows / total_rows if total_rows > 0 else 0
                iso_neg_pct[iso_name] = neg_pct
                log(f"    {iso_name}: {neg_rows}/{total_rows} negative ({neg_pct:.1%})")
            else:
                log(f"    {iso_name}: no data returned")
        except Exception as e:
            log(f"    {iso_name}: query failed ({e})")
            continue

    if not iso_neg_pct:
        log("WARNING: No negative LMP data retrieved from any ISO")
        return {}

    # Map ISO negative price frequency to counties via BA codes
    county_neg_freq: dict[str, float] = {}
    for ba_code, fips_set in ba_to_fips.items():
        # Find which ISO this BA belongs to
        iso_name = ba_to_iso.get(ba_code)
        if iso_name and iso_name in iso_neg_pct:
            for fips in fips_set:
                # Use max if county spans multiple BAs
                existing = county_neg_freq.get(fips, 0)
                county_neg_freq[fips] = max(existing, iso_neg_pct[iso_name])

    if not county_neg_freq:
        return {}

    # Normalize via percentile rank
    fips_list = list(county_neg_freq.keys())
    freq_values = [county_neg_freq[f] for f in fips_list]
    ranks = percentile_rank(freq_values)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    log(f"Negative LMP scores computed for {len(scores)} counties")
    return scores


# ============================================================
# Step 8: Interconnection Queue Pressure (LBNL "Queued Up")
# ============================================================

def compute_queue_pressure(tmpdir: str, fips_lookup: dict) -> dict[str, float]:
    """Compute interconnection queue pressure from LBNL Queued Up dataset.

    Downloads the LBNL interconnection queue dataset, parses active queue
    entries (renewable + storage projects), sums queued MW per county,
    and normalizes to 0-1 via percentile rank.

    Higher score = more MW in the interconnection queue = more developer
    activity and potential BTM opportunity for Nodiac.

    Returns: dict mapping FIPS -> queue pressure score (0-1).
    """
    print("\n=== Computing Queue Pressure (LBNL Queued Up) ===", flush=True)

    try:
        xlsx_data = download(LBNL_QUEUE_URL, "LBNL Queued Up (2025 Edition)")
    except Exception as e:
        log(f"WARNING: Failed to download LBNL queue data: {e}")
        return {}

    xlsx_path = os.path.join(tmpdir, "lbnl_queue.xlsx")
    with open(xlsx_path, "wb") as f:
        f.write(xlsx_data)

    log("Parsing LBNL queue data...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Find the main data sheet — usually "Data" or the first sheet
    data_sheet = None
    for sn in wb.sheetnames:
        sl = sn.lower()
        if sl in ("data", "queue data", "project data", "all projects"):
            data_sheet = sn
            break
    if data_sheet is None:
        # Try finding a sheet with many rows (the data sheet)
        # Fall back to the first non-codebook sheet
        for sn in wb.sheetnames:
            if "codebook" not in sn.lower() and "summary" not in sn.lower():
                data_sheet = sn
                break
    if data_sheet is None:
        data_sheet = wb.sheetnames[0]

    log(f"Using sheet: {data_sheet}")
    ws = wb[data_sheet]

    # Find header row
    headers = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip() if v else "" for v in row]
        # Look for key columns: state, county, capacity
        has_state = any("state" in v.lower() for v in vals)
        has_capacity = any("capacity" in v.lower() or "mw" in v.lower() for v in vals)
        if has_state and has_capacity:
            headers = vals
            header_row_idx = i
            break
        if i > 10:
            break

    if not headers:
        log("WARNING: Could not find header row in LBNL queue data")
        wb.close()
        return {}

    log(f"Header row found at index {header_row_idx}")
    log(f"Columns: {[h for h in headers if h][:15]}...")  # Show first 15 non-empty

    # Map column indices
    state_idx = None
    county_idx = None
    capacity_idx = None
    status_idx = None
    type_idx = None

    for j, h in enumerate(headers):
        hl = h.lower()
        if state_idx is None and ("state" in hl and "county" not in hl):
            state_idx = j
        elif county_idx is None and "county" in hl:
            county_idx = j
        elif capacity_idx is None and ("capacity" in hl or ("mw" in hl and "name" not in hl)):
            capacity_idx = j
        elif status_idx is None and "status" in hl:
            status_idx = j
        elif type_idx is None and ("type" in hl or "fuel" in hl or "technology" in hl):
            type_idx = j

    if state_idx is None or capacity_idx is None:
        log(f"WARNING: Missing required columns. state_idx={state_idx}, capacity_idx={capacity_idx}")
        wb.close()
        return {}

    log(f"Column indices: state={state_idx}, county={county_idx}, "
        f"capacity={capacity_idx}, status={status_idx}, type={type_idx}")

    # Active queue statuses (projects still seeking interconnection)
    ACTIVE_STATUSES = {
        "active", "pending", "in progress", "feasibility study",
        "system impact study", "facilities study", "ia pending",
        "ia executed", "under construction",
    }

    # Parse queue entries
    county_queued_mw: dict[str, float] = defaultdict(float)
    parsed = 0
    matched = 0
    skipped_status = 0
    skipped_no_location = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        vals = list(row)

        # Filter by status if available
        if status_idx is not None and status_idx < len(vals):
            status = str(vals[status_idx]).strip().lower() if vals[status_idx] else ""
            if status and status not in ACTIVE_STATUSES and "active" not in status:
                skipped_status += 1
                continue

        parsed += 1

        # Get capacity
        try:
            cap = float(vals[capacity_idx]) if vals[capacity_idx] else 0
        except (ValueError, TypeError):
            continue
        if cap <= 0:
            continue

        # Get state
        state = str(vals[state_idx]).strip() if state_idx is not None and state_idx < len(vals) and vals[state_idx] else ""
        if not state or state in SKIP_STATES:
            continue

        # Try to get county
        county = ""
        if county_idx is not None and county_idx < len(vals) and vals[county_idx]:
            county = str(vals[county_idx]).strip()

        if not county:
            # If no county, attribute to all counties in the state
            # by using state FIPS prefix
            skipped_no_location += 1
            continue

        fips = resolve_fips(state, county, fips_lookup)
        if fips:
            county_queued_mw[fips] += cap
            matched += 1

    wb.close()
    log(f"Queue entries parsed: {parsed}")
    log(f"  Matched to counties: {matched}")
    log(f"  Skipped (inactive status): {skipped_status}")
    log(f"  Skipped (no county): {skipped_no_location}")
    log(f"Counties with queued MW: {len(county_queued_mw)}")

    if not county_queued_mw:
        return {}

    # Log-normalize queued MW (very skewed distribution)
    log_mw = {f: math.log1p(mw) for f, mw in county_queued_mw.items()}

    # Percentile rank normalization
    fips_list = sorted(log_mw.keys())
    log_vals = [log_mw[f] for f in fips_list]
    ranks = percentile_rank(log_vals)

    scores: dict[str, float] = {}
    for fips, rank in zip(fips_list, ranks):
        scores[fips] = round(rank, 4)

    # Stats
    total_mw = sum(county_queued_mw.values())
    top = sorted(county_queued_mw.items(), key=lambda x: x[1], reverse=True)[:10]
    log(f"Total queued MW across all counties: {total_mw:,.0f}")
    log("Top 10 counties by queued MW:")
    for fips, mw in top:
        log(f"  {fips}: {mw:,.0f} MW (score={scores.get(fips, 0):.3f})")

    return scores


# ============================================================
# Step 9: Assemble final scores
# ============================================================

def assemble_scores(
    fips_lookup: dict,
    coop_scores: dict[str, float],
    grid_scores: dict[str, float],
    curtail_scores: dict[str, float],
    labor_scores: dict[str, float],
    fiber_scores: dict[str, float],
    queue_scores: dict[str, float],
    grid_metadata: dict[str, dict] | None = None,
    fiber_metadata: dict[str, dict] | None = None,
    fiber_source_map: dict[str, str] | None = None,
    ntia_metadata: dict[str, dict] | None = None,
) -> list[dict]:
    """Assemble all criterion scores into final county records."""
    print("\n=== Assembling Final Scores ===", flush=True)

    # Get base county list from Census FIPS file
    fips_data = download(FIPS_URL, "Census FIPS codes")
    text = fips_data.decode("utf-8")
    lines = text.strip().split("\n")

    counties = []

    for line in lines[1:]:  # Skip header
        parts = line.split("|")
        if len(parts) < 4:
            continue

        state_abbr = parts[0].strip()
        state_fips = parts[1].strip()
        county_fips_3 = parts[2].strip()
        county_name = parts[4].strip()

        if not state_fips or not county_fips_3:
            continue
        if state_abbr in SKIP_STATES or state_abbr == "DC":
            continue

        fips = state_fips + county_fips_3
        clean_name = county_name.replace(" County", "").replace(" Parish", "")

        # Determine data sources for this county
        sources: dict[str, str] = {}

        # Co-op density (area-based: % of county covered by co-op/public power territories)
        coop = coop_scores.get(fips)
        if coop is not None:
            sources["coop"] = "ArcGIS Co-op Service Territories (area-based, ORNL/LANL/INL)"
        else:
            coop = 0.0  # No territory overlap
            sources["coop"] = "Default (no co-op territory overlap)"

        # Grid reliability
        grid = grid_scores.get(fips)
        gm = (grid_metadata or {}).get(fips)
        if grid is not None and gm:
            year_range = f"{gm['min_year']}-{gm['max_year']}" if gm['min_year'] != gm['max_year'] else str(gm['min_year'])
            sources["grid"] = f"EIA Form 861 Reliability ({year_range}, {gm['year_count']}yr avg SAIDI)"
        elif grid is not None:
            sources["grid"] = "EIA Form 861 Reliability (actual)"
        else:
            # State-level fallback for grid (moderate default)
            grid = 0.5
            gm = None
            sources["grid"] = "Default estimate (no SAIDI data)"

        # Curtailment
        curtail = curtail_scores.get(fips)
        if curtail is not None:
            sources["curtail"] = (
                "EIA Forms 860 + 923 (plant-level CF gap analysis) + "
                "ISO/RTO curtailment intensity (CAISO, ERCOT, SPP, MISO, PJM 2023-2024 market reports)"
            )
        else:
            curtail = 0.0  # No renewables = no curtailment opportunity
            sources["curtail"] = "Default (no renewable generation)"

        # Permitting
        permitting = compute_permitting_score(state_abbr, fips)
        sources["permitting"] = (
            "State DC incentive programs (NCSL, SDI Alliance, H5 2025-2026), "
            "moratorium/opposition data (Data Center Watch 2025), "
            "county-level adjustments where data exists"
        )

        # Labor
        labor = labor_scores.get(fips)
        if labor is not None:
            sources["labor"] = "Census CBP 2023 (NAICS 5182/5415/517)"
        else:
            labor = 0.0
            sources["labor"] = "Default (no CBP data)"

        # Fiber
        fiber = fiber_scores.get(fips)
        if fiber is not None:
            sources["fiber"] = (fiber_source_map or {}).get(fips, "FCC BDC Dec 2024")
        else:
            fiber = 0.3  # Conservative default
            sources["fiber"] = "Default estimate (no fiber data)"

        # Queue pressure
        queue = queue_scores.get(fips)
        if queue is not None:
            sources["queue"] = "LBNL Queued Up 2025 Edition (interconnection queue MW through 2024)"
        else:
            queue = 0.0
            sources["queue"] = "Default (no queued projects)"

        county_record = {
            "fips_code": fips,
            "state_fips": state_fips,
            "county_name": clean_name,
            "state_abbr": state_abbr,
            "coop_density_score": round(coop, 4),
            "grid_reliability_score": round(grid, 4),
            "clipped_curtailed_score": round(curtail, 4),
            "permitting_score": round(permitting, 4),
            "labor_score": round(labor, 4),
            "fiber_score": round(fiber, 4),
            "queue_pressure_score": round(queue, 4),
            "data_sources": sources,
            "permitting_citation_ids": [],  # Populated by build_permitting_citations()
        }

        # Add fiber metadata if available (FCC BDC data)
        fm = (fiber_metadata or {}).get(fips)
        if fm:
            county_record["fiber_pct_availability"] = fm["pct_fiber"]
            county_record["fiber_bsls"] = fm["fiber_bsls"]
            county_record["fiber_total_bsls"] = fm["total_bsls"]
            county_record["fiber_providers"] = fm["fiber_providers"]

        # Add grid reliability metadata if available
        if gm:
            county_record["grid_reliability_years"] = gm["year_count"]
            county_record["grid_reliability_data_range"] = (
                f"{gm['min_year']}-{gm['max_year']}" if gm['min_year'] != gm['max_year']
                else str(gm['min_year'])
            )
            county_record["grid_reliability_avg_saidi"] = gm["avg_saidi"]

        # Add NTIA Middle Mile metadata if available
        nm = (ntia_metadata or {}).get(fips)
        if nm:
            county_record["ntia_middle_mile_grants"] = nm["grant_count"]
            county_record["ntia_middle_mile_amount"] = nm["total_amount"]
            county_record["ntia_middle_mile_recipients"] = nm["recipients"]

        counties.append(county_record)

    log(f"Total counties assembled: {len(counties)}")

    # Stats
    real_data_counts = {
        "coop": sum(1 for c in counties if "actual" in c["data_sources"]["coop"]),
        "grid": sum(1 for c in counties if "EIA" in c["data_sources"]["grid"]),
        "curtail": sum(1 for c in counties if "860" in c["data_sources"]["curtail"]),
        "labor": sum(1 for c in counties if "CBP" in c["data_sources"]["labor"]),
        "fiber": sum(1 for c in counties if "Default" not in c["data_sources"]["fiber"]),
        "queue": sum(1 for c in counties if "LBNL" in c["data_sources"].get("queue", "")),
    }
    for criterion, count in real_data_counts.items():
        pct = count / len(counties) * 100
        log(f"  {criterion}: {count}/{len(counties)} counties with real data ({pct:.0f}%)")

    return counties


# ============================================================
# Step 8: Write output
# ============================================================

def build_permitting_citations(counties: list[dict]) -> list[dict]:
    """Build permitting citation registry and assign IDs to each county.

    Returns the citation registry (list of {title, url, relevance} dicts).
    Mutates counties in place to set permitting_citation_ids.

    Citation architecture:
    - State-level citations: every county in a state gets its state's citations
    - County-level citations: specific counties with known opposition/moratoriums
    - Universal citations: methodology sources appended to all counties
    - Registry is deduplicated; counties reference by integer ID for compact JSON
    """
    print("\n=== Building Permitting Citations ===", flush=True)

    # Import citation data from the permitting module
    # (Defined inline to keep the pipeline self-contained)
    from _permitting_citations import STATE_CITS, COUNTY_CITS, UNIVERSAL_CITS

    registry: list[dict] = []
    key_to_id: dict[tuple[str, str], int] = {}

    def get_id(cit: dict) -> int:
        key = (cit["title"], cit["url"])
        if key not in key_to_id:
            key_to_id[key] = len(registry)
            registry.append(cit)
        return key_to_id[key]

    for county in counties:
        state = county["state_abbr"]
        fips = county["fips_code"]

        ids: list[int] = []
        for cit in STATE_CITS.get(state, []):
            ids.append(get_id(cit))
        for cit in COUNTY_CITS.get(fips, []):
            ids.append(get_id(cit))
        for cit in UNIVERSAL_CITS:
            ids.append(get_id(cit))

        # Deduplicate preserving order
        seen: set[int] = set()
        deduped: list[int] = []
        for i in ids:
            if i not in seen:
                seen.add(i)
                deduped.append(i)
        county["permitting_citation_ids"] = deduped

    log(f"Citation registry: {len(registry)} unique citations")
    avg = sum(len(c["permitting_citation_ids"]) for c in counties) / max(len(counties), 1)
    log(f"Average citations per county: {avg:.1f}")
    return registry


def write_output(counties: list[dict], citation_registry: list[dict] | None = None):
    """Write county scores to JSON file and optionally to Supabase."""
    print("\n=== Writing Output ===", flush=True)

    # Ensure output directory exists
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Wrap with citation registry if provided
    if citation_registry:
        output = {
            "permitting_citation_registry": citation_registry,
            "counties": counties,
        }
    else:
        output = counties

    # Write static JSON (compact for smaller file size)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, separators=(",", ":"))

    file_size = OUTPUT_PATH.stat().st_size / 1024
    log(f"Wrote {OUTPUT_PATH} ({file_size:.0f} KB, {len(counties)} counties)")

    # Optional Supabase upsert
    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_SECRET_KEY") or os.environ.get("NEXT_PUBLIC_SUPABASE_PUBLISHABLE_KEY")

    if supabase_url and supabase_key:
        log("Upserting to Supabase...")
        try:
            from supabase import create_client
            client = create_client(supabase_url, supabase_key)

            batch_size = 500
            inserted = 0
            for i in range(0, len(counties), batch_size):
                batch = counties[i:i + batch_size]
                result = client.table("county_scores").upsert(
                    batch, on_conflict="fips_code"
                ).execute()
                inserted += len(batch)
                log(f"  Batch {i // batch_size + 1}: {inserted}/{len(counties)}")

            log(f"Supabase upsert complete: {inserted} counties")
        except ImportError:
            log("WARNING: supabase-py not installed, skipping DB upsert")
            log("  Run: uv add supabase to enable")
        except Exception as e:
            log(f"WARNING: Supabase upsert failed: {e}")
    else:
        log("Supabase credentials not found, skipping DB upsert")
        log("  Set NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SECRET_KEY to enable")


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 60)
    print("  NODIAC REGIONAL HUB STRATEGY — REAL DATA PIPELINE")
    print("=" * 60)

    with tempfile.TemporaryDirectory(prefix="nodiac-data-") as tmpdir:
        log(f"Temp directory: {tmpdir}")

        # Build FIPS lookup
        fips_lookup = build_fips_lookup()

        # Compute each criterion score
        coop_scores = compute_coop_density_area(fips_lookup)
        grid_scores, grid_metadata = compute_grid_reliability(tmpdir, fips_lookup)
        curtail_scores, plant_info = compute_curtailment_score(tmpdir, fips_lookup)
        labor_scores_raw = compute_labor_score()
        labor_scores = blend_labor_with_neighbors(labor_scores_raw)

        # NTIA Middle Mile grants (computed before fiber so it can be blended in)
        ntia_scores, ntia_metadata = compute_ntia_middle_mile(fips_lookup)

        fiber_scores, fiber_metadata, fiber_source_map = compute_fiber_score(
            ntia_scores=ntia_scores,
        )
        queue_scores = compute_queue_pressure(tmpdir, fips_lookup)

        # Enhance curtailment with negative LMP frequency (optional)
        neg_lmp_scores = compute_negative_lmp_frequency(plant_info, fips_lookup)
        if neg_lmp_scores:
            log(f"\nIntegrating negative LMP data into curtailment scores ({len(neg_lmp_scores)} counties)...")
            for fips in curtail_scores:
                if fips in neg_lmp_scores:
                    # Blend: 85% existing curtailment + 15% negative LMP frequency
                    curtail_scores[fips] = round(
                        0.85 * curtail_scores[fips] + 0.15 * neg_lmp_scores[fips],
                        4,
                    )

        # Assemble final scores
        counties = assemble_scores(
            fips_lookup,
            coop_scores,
            grid_scores,
            curtail_scores,
            labor_scores,
            fiber_scores,
            queue_scores,
            grid_metadata,
            fiber_metadata,
            fiber_source_map,
            ntia_metadata,
        )

        # Build permitting citations
        try:
            citation_registry = build_permitting_citations(counties)
        except ImportError:
            log("WARNING: _permitting_citations module not found, skipping citations")
            log("  Citations will be empty. See scripts/_permitting_citations.py")
            citation_registry = None

        # Write output
        write_output(counties, citation_registry)

    print("\n" + "=" * 60)
    print("  PIPELINE COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
